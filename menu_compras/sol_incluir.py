import sys
from banco_dados.conexao import conecta
from comandos.comando_notificacao import mensagem_alerta, pergunta_confirmacao, tratar_notificar_erros
from comandos.comando_tabelas import extrair_tabela, lanca_tabela, limpa_tabela, layout_cabec_tab, excluir_item_tab
from comandos.comando_lines import definir_data_atual
from comandos.comando_cores import cor_amarelo, cor_branco, cor_vermelho
from comandos.comando_telas import tamanho_aplicacao, icone, cor_widget, cor_widget_cab, cor_fonte, cor_btn
from comandos.comando_telas import cor_fundo_tela
from comandos.comando_banco import definir_proximo_generator
from comandos.comando_conversoes import valores_para_float
from banco_dados.bc_consultas import Produto, ProdutoOrdemSolicitacao, ProdutoOrdemRequisicao, ProdutoOrdemCompra
from banco_dados.bc_consultas import Projeto, MateriaPrima
from forms.tela_sol_incluir import *
from PyQt5.QtWidgets import QApplication, QFileDialog, QShortcut, QMainWindow
from PyQt5.QtGui import QKeySequence, QFont, QColor
from PyQt5.QtCore import Qt
from datetime import date, datetime
from unidecode import unidecode
import os
import shutil
import socket
import math
import inspect
from functools import partial

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Side, Alignment, Border, Font
import openpyxl.styles as styles
from pathlib import Path


class TelaSolIncluir(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        cor_fundo_tela(self)
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_compra_sol.png")
        tamanho_aplicacao(self)
        self.layout_tabela_recomendacao(self.table_Recomendacao)
        self.layout_tabela_solicitacao(self.table_Solicitacao)
        self.layout_tabela_anexos(self.table_Anexos)
        self.layout_proprio()

        self.tab_prod = Produto()
        self.tab_prod_sol = ProdutoOrdemSolicitacao()
        self.tab_prod_req = ProdutoOrdemRequisicao()
        self.tab_prod_oc = ProdutoOrdemCompra()
        self.tab_projeto = Projeto()
        self.tab_materiaprima = MateriaPrima()

        self.tab_shortcut = QShortcut(QKeySequence(Qt.Key_Tab), self)
        self.tab_shortcut.activated.connect(self.manipula_tab)

        self.table_Recomendacao.viewport().installEventFilter(self)

        self.definir_botoes_e_comandos()
        self.reiniciando_sol()

        self.processando = False

    def layout_proprio(self):
        try:
            cor_widget_cab(self.widget_cabecalho)

            cor_widget(self.widget_Cor1)
            cor_widget(self.widget_Cor2)
            cor_widget(self.widget_Cor3)
            cor_widget(self.widget_Cor4)
            cor_widget(self.widget_Cor5)
            cor_widget(self.widget_Cor6)
            cor_widget(self.widget_Cor7)
            cor_widget(self.widget_Cor8)

            cor_fonte(self.label)
            cor_fonte(self.label_11)
            cor_fonte(self.label_10)
            cor_fonte(self.label_12)
            cor_fonte(self.label_13)
            cor_fonte(self.label_14)
            cor_fonte(self.label_15)
            cor_fonte(self.label_2)
            cor_fonte(self.label_28)
            cor_fonte(self.label_23)
            cor_fonte(self.label_29)
            cor_fonte(self.label_25)
            cor_fonte(self.label_3)
            cor_fonte(self.label_37)
            cor_fonte(self.label_32)
            cor_fonte(self.label_33)
            cor_fonte(self.label_35)
            cor_fonte(self.label_36)
            cor_fonte(self.label_Referencia)
            cor_fonte(self.label_40)
            cor_fonte(self.label_41)
            cor_fonte(self.label_42)
            cor_fonte(self.label_45)
            cor_fonte(self.label_49)
            cor_fonte(self.label_5)
            cor_fonte(self.label_53)
            cor_fonte(self.label_52)
            cor_fonte(self.label_54)
            cor_fonte(self.label_57)
            cor_fonte(self.label_58)
            cor_fonte(self.label_59)
            cor_fonte(self.label_6)
            cor_fonte(self.label_7)
            cor_fonte(self.label_8)
            cor_fonte(self.label_9)
            cor_fonte(self.label_Titulo)
            cor_fonte(self.check_Nivel)
            cor_fonte(self.check_Excel)

            cor_btn(self.btn_Salvar)
            cor_btn(self.btn_ExcluirTudo_Sol)
            cor_btn(self.btn_ExcluirItem_Sol)
            cor_btn(self.btn_ExcluirItem_Rec)
            cor_btn(self.btn_Adicionar_Chapas)
            cor_btn(self.btn_Adicionar_Todos)
            cor_btn(self.btn_Adicionar_SemSaldo)
            cor_btn(self.btn_ExcluirTudo_Rec)
            cor_btn(self.btn_Consultar_Estrut)
            cor_btn(self.btn_Consome_Manu)
            cor_btn(self.btn_Consultar_Consumo)
            cor_btn(self.btn_SelecionarAnexo)
            cor_btn(self.btn_ExcluirAnexo)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def calcula_embalagem_sim(self, codigo_produto, qtde):
        try:
            qtde_final = None
            texto = None

            dados_prod = self.tab_prod.consulta_por_codigo(codigo_produto)
            um = dados_prod[0][4]
            embalagem = dados_prod[0][7]
            kg_mt = dados_prod[0][8]

            if embalagem == "SIM":
                if um == "KG":
                    metros = float(qtde) / float(kg_mt)
                    barra_quebra = metros / 6

                    barra_arred = math.ceil(barra_quebra)
                    mts_totais = barra_arred * 6

                    texto = f"{barra_arred} BARRAS DE 6 MTS"

                    total_kg = mts_totais * float(kg_mt)
                    qtde_final = "%.2f" % total_kg

            return qtde_final, texto

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def soma_e_classifica(self, dados):
        try:
            produto_dict = {}

            for produto in dados:
                codigo = produto[0]
                quantidade = produto[4]

                if codigo in produto_dict:
                    produto_dict[codigo] += quantidade
                else:
                    produto_dict[codigo] = quantidade

            novo_produto_lista = []
            for codigo, quantidade in produto_dict.items():
                for produto in dados:
                    if produto[0] == codigo:
                        novo_produto_lista.append((produto[0], produto[1], produto[2], produto[3], quantidade,
                                                   produto[5], produto[6]))
                        break

            lista_de_listas_ordenada = sorted(novo_produto_lista, key=lambda x: x[1])
            return lista_de_listas_ordenada

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def consulta_compras_pendentes(self, codigo_produto):
        try:
            dados_prod = self.tab_prod.consulta_por_codigo(codigo_produto)
            id_prod = dados_prod[0][0]
            descr = dados_prod[0][1]
            um = dados_prod[0][4]

            dados = []

            select_solicitacao = self.tab_prod_sol.consulta_por_produto_aberto(id_prod)
            if select_solicitacao:
                for sol in select_solicitacao:
                    mestre_sol = sol[0]
                    qtde_sol = sol[1]
                    data_sol = sol[2]

                    data_sol_e = '{}/{}/{}'.format(data_sol.day, data_sol.month, data_sol.year)

                    info_sol = f"Solicitação Nº {mestre_sol} - {qtde_sol} {um} - Emissão: {data_sol_e}"
                    dados.append(info_sol)

            select_requisicao = self.tab_prod_req.consulta_por_produto_aberto(id_prod)
            if select_requisicao:
                for req in select_requisicao:
                    num_req = req[0]
                    qtde_req = req[1]
                    data_req = req[2]

                    data_req_e = '{}/{}/{}'.format(data_req.day, data_req.month, data_req.year)

                    info_req = f"Requisição Nº {num_req} - {qtde_req} {um} - Emissão: {data_req_e}"
                    dados.append(info_req)

            select_oc = self.tab_prod_oc.consulta_por_produto_aberto(id_prod)
            if select_oc:
                for oc in select_oc:
                    num_oc = oc[0]
                    qtde_oc = oc[1]
                    data_oc = oc[2]

                    data_oc_e = '{}/{}/{}'.format(data_oc.day, data_oc.month, data_oc.year)

                    info_oc = f"Ordem de Compra Nº {num_oc} - {qtde_oc} {um} - Emissão: {data_oc_e}"
                    dados.append(info_oc)
            if dados:
                if len(dados) == 1:
                    msg = f'Existem solicitações de compra em aberto do item\n\n' \
                          f'Cód. {codigo_produto} - {descr}:\n\n' \
                          f'  - {dados[0]}\n\n' \
                          f'Deseja continuar?'
                    if pergunta_confirmacao(msg):
                        status = True
                    else:
                        status = False

                else:
                    tem = ''
                    for valor in dados:
                        tem = tem + "- " + valor + "\n"

                    msg1 = f'Existem solicitações de compra em aberto do item\n\n' \
                           f'Cód. {codigo_produto} - {descr}:\n\n' \
                           f'{tem}\n\n' \
                           f'Deseja continuar?'
                    if pergunta_confirmacao(msg1):
                        status = True
                    else:
                        status = False

            else:
                status = True

            return status

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def layout_tabela_recomendacao(self, nome_tabela):
        try:
            layout_cabec_tab(nome_tabela)

            nome_tabela.setColumnWidth(0, 35)
            nome_tabela.setColumnWidth(1, 200)
            nome_tabela.setColumnWidth(2, 85)
            nome_tabela.setColumnWidth(3, 35)
            nome_tabela.setColumnWidth(4, 45)
            nome_tabela.setColumnWidth(5, 95)
            nome_tabela.setColumnWidth(6, 50)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def layout_tabela_solicitacao(self, nome_tabela):
        try:
            layout_cabec_tab(nome_tabela)

            nome_tabela.setColumnWidth(0, 35)
            nome_tabela.setColumnWidth(1, 160)
            nome_tabela.setColumnWidth(2, 80)
            nome_tabela.setColumnWidth(3, 35)
            nome_tabela.setColumnWidth(4, 45)
            nome_tabela.setColumnWidth(5, 130)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def layout_tabela_anexos(self, nome_tabela):
        try:
            layout_cabec_tab(nome_tabela)

            nome_tabela.setColumnWidth(0, 150)
            nome_tabela.setColumnWidth(1, 300)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def manipula_tab(self):
        try:
            if self.line_Codigo_Estrut.hasFocus():
                self.verifica_line_codigo()

            elif self.line_Qtde_Estrut.hasFocus():
                self.verifica_line_qtde()

            elif self.line_Codigo_Manu.hasFocus():
                self.verifica_line_codigo_manu()

            elif self.line_Destino_Manu.hasFocus():
                self.verifica_line_qtde_manu()

            elif self.line_Qtde_Manu.hasFocus():
                self.line_Destino_Manu.setFocus()

            elif self.line_Unidade.hasFocus():
                self.combo_Unidade.setFocus()

            elif self.combo_Unidade.hasFocus():
                self.line_Medida.setFocus()

            elif self.line_Medida.hasFocus():
                self.combo_Medida.setFocus()

            elif self.combo_Medida.hasFocus():
                self.line_Destino_Manu.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def eventFilter(self, source, event):
        try:
            if (event.type() == QtCore.QEvent.MouseButtonDblClick and
                    event.buttons() == QtCore.Qt.LeftButton and
                    source is self.table_Recomendacao.viewport()):

                item = self.table_Recomendacao.currentItem()

                destino = ''
                desc_pai = self.line_Descricao_Estrut.text()
                origem = self.combo_Consumo.currentText()

                if desc_pai:
                    destino = desc_pai

                elif origem:
                    origemtete = origem.find(" - ")
                    nome_origem = origem[origemtete:]

                    destino = f"CI{nome_origem}"

                extrai_recomendados = extrair_tabela(self.table_Recomendacao)
                item_selecionado = extrai_recomendados[item.row()]

                cod, desc, ref, um, qtde, local, saldo = item_selecionado
                quanti, texto = self.calcula_embalagem_sim(cod, qtde)

                if quanti:
                    qtde = quanti
                if texto:
                    ref = texto

                dados_prod = self.tab_prod.consulta_por_codigo(cod)
                ncm = dados_prod[0][9]
                if ncm:
                    existe_compra = self.consulta_compras_pendentes(cod)
                    if existe_compra:
                        extrai_solicitacao = extrair_tabela(self.table_Solicitacao)

                        ja_existe = False
                        for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                            if c_sol == cod:
                                ja_existe = True
                                break

                        if not ja_existe:
                            didis = [cod, desc, ref, um, float(qtde), destino]
                            extrai_solicitacao.append(didis)

                            self.procedimento_lanca_tabela(extrai_solicitacao)
                        else:
                            mensagem_alerta(f'O item selecionado já está presente na tabela'
                                            f'"Lista Solicitação".')
                            didis = [cod, desc, ref, um, float(qtde), destino]
                            extrai_solicitacao.append(didis)

                            self.procedimento_lanca_tabela(extrai_solicitacao)
                else:
                    mensagem_alerta(f'Este produto está sem "NCM" no cadastro.\n\n'
                                    f'Aproveite para verificar se o produto está apto '
                                    f'para compra no Siger.')

            return super(QMainWindow, self).eventFilter(source, event)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def definir_lineedit_bloqueados(self):
        try:
            self.line_Descricao_Manu.setReadOnly(True)
            self.line_Referencia_Manu.setReadOnly(True)
            self.line_Local_Manu.setReadOnly(True)
            self.line_UM_Manu.setReadOnly(True)
            self.line_Saldo_Manu.setReadOnly(True)

            self.line_Descricao_Estrut.setReadOnly(True)
            self.line_Referencia_Estrut.setReadOnly(True)
            self.line_Local_Estrut.setReadOnly(True)
            self.line_UM_Estrut.setReadOnly(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def desaparece_referencia_editada(self):
        try:
            self.widget_Kilos.setHidden(True)
            self.widget_Referencia.setHidden(False)
            self.label_Referencia.setText("Referência:")
            self.label_Medida.setText("")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def aparece_referencia_editada(self):
        try:
            self.widget_Kilos.setHidden(False)
            self.widget_Referencia.setHidden(True)
            self.label_Referencia.setText("Qtde Barras/Peças:")
            self.label_Referencia.setMinimumWidth(120)
            self.label_Referencia.setMaximumWidth(120)
            self.label_Medida.setText("Medida:")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def definir_combo_consumo(self):
        try:
            self.combo_Consumo.clear()

            nova_lista = [""]

            projetos = self.tab_projeto.consulta_todos_itens()
            for ides, descr in projetos:
                dd = f"{ides} - {descr}"
                nova_lista.append(dd)

            self.combo_Consumo.addItems(nova_lista)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def definir_line_ano_consumo(self):
        try:
            data_hoje = date.today()
            ano_atual = data_hoje.strftime("%Y")
            ano_menos_dois = str(int(ano_atual) - 2)
            self.line_Ano_Consumo.setText(ano_menos_dois)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def definir_botoes_e_comandos(self):
        try:
            self.line_Codigo_Manu.editingFinished.connect(self.verifica_line_codigo_manu)

            self.line_Codigo_Estrut.returnPressed.connect(lambda: self.verifica_line_codigo())

            self.btn_Consultar_Estrut.clicked.connect(self.verifica_line_qtde)
            self.line_Qtde_Estrut.returnPressed.connect(lambda: self.verifica_line_qtde())

            self.btn_Consome_Manu.clicked.connect(self.verifica_line_qtde_manu)
            self.line_Destino_Manu.returnPressed.connect(lambda: self.verifica_line_qtde_manu())

            self.line_Qtde_Manu.returnPressed.connect(lambda: self.line_Destino_Manu.setFocus())

            self.btn_Consultar_Consumo.clicked.connect(self.procura_origem)

            self.btn_SelecionarAnexo.clicked.connect(self.procura_anexo)

            self.btn_ExcluirAnexo.clicked.connect(partial(excluir_item_tab, self.table_Anexos, "Lista Anexos"))

            self.btn_ExcluirTudo_Sol.clicked.connect(partial(limpa_tabela, self.table_Solicitacao))
            self.btn_ExcluirItem_Sol.clicked.connect(partial(excluir_item_tab, self.table_Solicitacao,
                                                             "Lista Solicitação"))

            self.btn_ExcluirTudo_Rec.clicked.connect(partial(limpa_tabela, self.table_Recomendacao))
            self.btn_ExcluirItem_Rec.clicked.connect(partial(excluir_item_tab, self.table_Recomendacao,
                                                             "Lista de Recomendações"))

            self.btn_Adicionar_Todos.clicked.connect(self.tudo_verifica_deondevem)

            self.btn_Adicionar_SemSaldo.clicked.connect(self.tudo_verifica_deondevem)

            self.btn_Adicionar_Chapas.clicked.connect(self.lanca_so_chapas)

            self.btn_Salvar.clicked.connect(self.verifica_salvamento)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def procura_anexo(self):
        try:
            extrai_anexos = extrair_tabela(self.table_Anexos)

            file_dialog = QFileDialog()
            file_dialog.setFileMode(QFileDialog.ExistingFiles)
            desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
            file_paths, _ = file_dialog.getOpenFileNames(None, 'Selecionar Anexo', directory=desktop_path)

            if file_paths:
                for file_path in file_paths:
                    nome_arquivo = os.path.basename(file_path)

                    ja_existe = False
                    for arq_ext, cam_ext in extrai_anexos:
                        if arq_ext == nome_arquivo:
                            ja_existe = True
                            break

                    if not ja_existe:
                        dados = [nome_arquivo, file_path]
                        extrai_anexos.append(dados)
                    else:
                        mensagem_alerta(f'O arquivo {nome_arquivo} já foi adicionado!!')

            if extrai_anexos:
                lanca_tabela(self.table_Anexos, extrai_anexos)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def definir_validador_lineedit(self):
        try:
            validator = QtGui.QDoubleValidator(0, 9999999.000, 3, self.line_Qtde_Manu)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Qtde_Manu.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Codigo_Estrut)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Codigo_Estrut.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Num_Sol)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Num_Sol.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Codigo_Manu)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Codigo_Manu.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456, self.line_Unidade)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Unidade.setValidator(validator)

            validator = QtGui.QIntValidator(0, 123456789, self.line_Medida)
            locale = QtCore.QLocale("pt_BR")
            validator.setLocale(locale)
            self.line_Medida.setValidator(validator)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def procura_origem(self):
        try:
            self.reiniciando_produto_estrutura()
            self.table_Recomendacao.setRowCount(0)

            origem = self.combo_Consumo.currentText()
            if origem:
                origemtete = origem.find(" - ")
                id_origem = origem[:origemtete]
                ano = self.line_Ano_Consumo.text()

                compras = self.tab_projeto.consulta_por_codigo_data(id_origem, ano)
                if not compras:
                    mensagem_alerta('Não encontramos histórico de compras neste período')
                    self.table_Recomendacao.clearContents()
                else:
                    lanca_tabela(self.table_Recomendacao, compras)
                    self.pintar_tabela_recomendacao()
            else:
                mensagem_alerta(f'Defina um "Setor"!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def verifica_line_codigo(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()
            if len(codigo_produto) == 0:
                mensagem_alerta('O campo "Código" não pode estar vazio')
                self.line_Codigo_Estrut.clear()
            elif int(codigo_produto) == 0:
                mensagem_alerta('O campo "Código" não pode ser "0"')
                self.line_Codigo_Estrut.clear()
            else:
                self.verifica_sql_codigo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def verifica_sql_codigo(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()

            detalhes_produto = self.tab_prod.consulta_por_codigo(codigo_produto)

            produto_acabado = self.tab_prod.verifica_acabado(codigo_produto)

            if not detalhes_produto:
                mensagem_alerta('Este código de produto não existe!')
                self.line_Codigo_Estrut.clear()
            elif not produto_acabado:
                mensagem_alerta('Este código não está classificado como "Produto Acabado"!')
                self.line_Codigo_Estrut.clear()
            else:
                self.lanca_dados_codigo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lanca_dados_codigo(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()

            dados_prod = self.tab_prod.consulta_por_codigo(codigo_produto)
            descr = dados_prod[0][1]
            ref = dados_prod[0][3]
            if ref:
                referencia = ref
            else:
                referencia = ""
            um = dados_prod[0][4]
            saldo = dados_prod[0][6]

            saldo_float = valores_para_float(saldo)

            if saldo_float < 0:
                mensagem_alerta(f'Este produto está com saldo negativo!\n'
                                f'Saldo Total = {saldo_float}')
                self.line_Codigo_Estrut.clear()
            else:
                self.line_Descricao_Estrut.setText(descr)
                self.line_Referencia_Estrut.setText(referencia)
                self.line_UM_Estrut.setText(um)
                self.line_Qtde_Estrut.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def verifica_line_qtde(self):
        try:
            self.combo_Consumo.setCurrentText("")

            qtde = self.line_Qtde_Estrut.text()
            if not qtde:
                mensagem_alerta('O campo "Qtde:" não pode estar vazio')
                self.line_Qtde_Estrut.clear()
                self.line_Qtde_Estrut.setFocus()
            else:
                qtde_float = valores_para_float(qtde)

                if qtde_float == 0:
                    mensagem_alerta('O campo "Qtde:" não pode ser "0"')
                    self.line_Qtde_Estrut.clear()
                    self.line_Qtde_Estrut.setFocus()
                else:
                    if self.check_Nivel.isChecked():
                        self.lanca_todos_niveis()
                    else:
                        self.lanca_estrutura()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lanca_estrutura(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()
            qtde = self.line_Qtde_Estrut.text()

            qtde_float = valores_para_float(qtde)

            dados_prod = self.tab_prod.consulta_por_codigo(codigo_produto)
            id_prod = dados_prod[0][0]

            dados_estrutura = self.tab_materiaprima.consulta_estrutura_pro_produto(id_prod, qtde_float)

            nova_tabela = []

            if not dados_estrutura:
                mensagem_alerta(f'Este produto não possui estrutura cadastrada!\n'
                                f'Antes de criar a Ordem de Produção, defina a estrutura.')
                self.reiniciando_produto_estrutura()

            else:
                for dados in dados_estrutura:
                    cod = dados[0]
                    descr = dados[1]
                    ref = dados[2]
                    if ref:
                        referencia = ref
                    else:
                        referencia = ""
                    um = dados[3]
                    qtde = dados[4]
                    local = dados[5]
                    saldo = dados[6]

                    dados_prod = self.tab_prod.consulta_por_codigo(cod)
                    conjunto = dados_prod[0][10]
                    terceirizado = dados_prod[0][11]

                    if conjunto == 10:
                        if terceirizado:
                            didos = (cod, descr, referencia, um, qtde, local, saldo)
                            nova_tabela.append(didos)
                    else:
                        didos = (cod, descr, referencia, um, qtde, local, saldo)
                        nova_tabela.append(didos)

            if nova_tabela:
                lanca_tabela(self.table_Recomendacao, nova_tabela)
                self.pintar_tabela_recomendacao()
            else:
                mensagem_alerta(f'Este produto não possui material comprado na estrutura!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lanca_todos_niveis(self):
        try:
            codigo_produto = self.line_Codigo_Estrut.text()
            qtde = self.line_Qtde_Estrut.text()

            qtde_float = valores_para_float(qtde)

            tudo_tudo = []
            estrutura = self.verifica_estrutura(codigo_produto, qtde_float)
            for kuku in estrutura:
                tudo_tudo.append(kuku)

            tabela_estrutura = self.soma_e_classifica(tudo_tudo)

            nova_tabela = []
            if not tabela_estrutura:
                mensagem_alerta(f'Este produto não possui estrutura cadastrada!\n'
                                f'Antes de criar a Ordem de Produção, defina a estrutura.')
                self.reiniciando_produto_estrutura()

            else:
                for dados in tabela_estrutura:
                    cod, descr, ref, um, qtde, local, saldo = dados

                    dados_prod = self.tab_prod.consulta_por_codigo(cod)
                    conjunto = dados_prod[0][10]
                    terceirizado = dados_prod[0][11]

                    if conjunto == 10:
                        if terceirizado:
                            didos = (cod, descr, ref, um, qtde, local, saldo)
                            nova_tabela.append(didos)
                    else:
                        didos = (cod, descr, ref, um, qtde, local, saldo)
                        nova_tabela.append(didos)

            if nova_tabela:
                lanca_tabela(self.table_Recomendacao, nova_tabela)
                self.pintar_tabela_recomendacao()
            else:
                mensagem_alerta(f'Este produto não possui material comprado na estrutura!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def verifica_estrutura(self, codigos, qtde):
        try:
            dados_prod = self.tab_prod.consulta_por_codigo(codigos)
            id_prod = dados_prod[0][0]
            descr = dados_prod[0][1]
            ref = dados_prod[0][3]
            um = dados_prod[0][4]
            local = dados_prod[0][5]
            saldo = dados_prod[0][6]

            dadoss = (codigos, descr, ref, um, qtde, local, saldo)
            filhos = [dadoss]

            estrutura_filho = self.tab_materiaprima.consulta_estrutura_pro_produto(id_prod, qtde)

            for dados_f in estrutura_filho:
                cod_f = dados_f[0]
                qtde_f = dados_f[4]

                filhos.extend(self.verifica_estrutura(cod_f, qtde_f))

            return filhos

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def reiniciando_produto_estrutura(self):
        try:
            self.line_Codigo_Estrut.clear()
            self.line_Descricao_Estrut.clear()
            self.line_Referencia_Estrut.clear()
            self.line_UM_Estrut.clear()
            self.line_Qtde_Estrut.clear()
            self.line_Codigo_Estrut.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def pintar_tabela_recomendacao(self):
        try:
            extrai_tabela = extrair_tabela(self.table_Recomendacao)

            for index, itens in enumerate(extrai_tabela):
                cod, descr, ref, um, qtde, local, saldo = itens

                qtde_float = valores_para_float(qtde)
                saldo_float = valores_para_float(saldo)

                if saldo_float < qtde_float:
                    font = QFont()
                    font.setBold(True)

                    self.table_Recomendacao.item(index, 6).setBackground(QColor(cor_vermelho))
                    self.table_Recomendacao.item(index, 6).setFont(font)
                    self.table_Recomendacao.item(index, 6).setForeground(QColor(cor_branco))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def pintar_bloquear_tabela_solicitacao(self):
        try:
            dados_tabela = extrair_tabela(self.table_Solicitacao)

            fonte = QFont("Segoe UI", 7)

            for index, dados in enumerate(dados_tabela):
                cod, descr, ref, um, qtde, destino = dados

                dados_prod = self.tab_prod.consulta_por_codigo(cod)
                embalagem = dados_prod[0][7]

                if cod == "1":
                    self.table_Solicitacao.item(index, 1).setBackground(QColor(cor_amarelo))
                    self.table_Solicitacao.item(index, 2).setBackground(QColor(cor_amarelo))
                    self.table_Solicitacao.item(index, 3).setBackground(QColor(cor_amarelo))

                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][0]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 0, item)

                elif embalagem == "SIM" or embalagem == "SER":
                    self.table_Solicitacao.item(index, 2).setBackground(QColor(cor_amarelo))

                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][0]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 0, item)

                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][1]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 1, item)

                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][3]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 3, item)

                else:
                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][0]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 0, item)

                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][1]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 1, item)

                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][2]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 2, item)

                    item = QtWidgets.QTableWidgetItem(str(dados_tabela[index][3]))
                    item.setFlags(QtCore.Qt.ItemIsEnabled)
                    self.table_Solicitacao.setItem(index, 3, item)

                self.table_Solicitacao.item(index, 1).setFont(fonte)
                self.table_Solicitacao.item(index, 2).setFont(fonte)

            for row in range(self.table_Solicitacao.rowCount()):
                if row % 2 == 0:
                    for col in range(self.table_Solicitacao.columnCount()):
                        item = self.table_Solicitacao.item(row, col)
                        item.setBackground(QColor(220, 220, 220))

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def verifica_line_codigo_manu(self):
        if not self.processando:
            try:
                self.processando = True

                codigo_produto = self.line_Codigo_Manu.text()

                if codigo_produto:
                    if int(codigo_produto) == 0:
                        mensagem_alerta('O campo "Código" não pode ser "0"')
                        self.line_Codigo_Manu.clear()
                    else:
                        self.verifica_sql_produtomanual()

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

            finally:
                self.processando = False

    def verifica_sql_produtomanual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()

            dados_prod = self.tab_prod.consulta_por_codigo(codigo_produto)

            if not dados_prod:
                mensagem_alerta('Este código de produto não existe!')
                self.line_Codigo_Manu.clear()
            else:
                self.verifica_materia_prima()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def verifica_materia_prima(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()

            dados_prod = self.tab_prod.consulta_por_codigo(codigo_produto)
            conjunto = dados_prod[0][10]
            terceirizado = dados_prod[0][11]
            tipo = dados_prod[0][12]

            if conjunto == 10:
                if terceirizado and tipo == 119:
                    self.lanca_dados_produtomanual()
                else:
                    msg_produto = 'Materiais definidos como "Produtos Acabados" precisam ter:\n\n' \
                                  '- O custo do serviço vinculado a Estrutura.\n\n' \
                                  '- O "Tipo de Material" cadastrado como "INDUSTRIALIZACAO."'
                    mensagem_alerta(f'{msg_produto}')
                    self.line_Codigo_Manu.clear()

            else:
                self.lanca_dados_produtomanual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lanca_dados_produtomanual(self):
        try:
            codigo_produto = self.line_Codigo_Manu.text()

            dados_prod = self.tab_prod.consulta_por_codigo(codigo_produto)
            descr = dados_prod[0][1]
            ref = dados_prod[0][3]
            um = dados_prod[0][4]
            local = dados_prod[0][5]
            saldo = dados_prod[0][6]
            embalagem = dados_prod[0][7]
            kg_mt = dados_prod[0][8]

            self.line_Descricao_Manu.setText(descr)
            self.line_UM_Manu.setText(um)
            self.line_Local_Manu.setText(local)
            numero = str(saldo).replace('.', ',')
            self.line_Saldo_Manu.setText(numero)
            self.line_Qtde_Manu.setEnabled(True)
            self.btn_Consome_Manu.setEnabled(True)

            if embalagem == "SIM":
                if um == "KG" and not kg_mt:
                    mensagem_alerta(f'Defina, no cadastro do produto, quantos kg tem 1 metro de material!')
                else:
                    self.aparece_referencia_editada()
                    self.line_Unidade.setFocus()
                    self.setTabOrder(self.line_Unidade, self.combo_Unidade)
                    self.setTabOrder(self.combo_Unidade, self.line_Medida)
                    self.setTabOrder(self.line_Medida, self.combo_Medida)
                    self.setTabOrder(self.combo_Medida, self.line_Qtde_Manu)
                    self.setTabOrder(self.line_Qtde_Manu, self.line_Destino_Manu)
            elif embalagem == "SER":
                self.desaparece_referencia_editada()
                self.line_Referencia_Manu.setReadOnly(False)
                self.line_Referencia_Manu.setText(ref)
                self.line_Referencia_Manu.setFocus()
                self.setTabOrder(self.line_Referencia_Manu, self.line_Qtde_Manu)
                self.setTabOrder(self.line_Qtde_Manu, self.line_Destino_Manu)
            else:
                self.desaparece_referencia_editada()
                self.line_Referencia_Manu.setText(ref)
                self.line_Qtde_Manu.setFocus()
                self.setTabOrder(self.line_Qtde_Manu, self.line_Destino_Manu)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def verifica_line_qtde_manu(self):
        try:
            codiguzinho = self.line_Codigo_Manu.text()
            codigo_produto = int(codiguzinho)

            line_unidade = self.line_Unidade.text()
            combo_unidade = self.combo_Unidade.currentText()

            line_medida = self.line_Medida.text()
            combo_medida = self.combo_Medida.currentText()

            dados_prod = self.tab_prod.consulta_por_codigo(codigo_produto)
            embalagem = dados_prod[0][7]

            if embalagem == "SIM":
                if not line_unidade or not combo_unidade:
                    mensagem_alerta('Informe a unidade de medida no campo "Referência"')
                    self.line_Unidade.setFocus()
                elif not line_medida or not combo_medida:
                    mensagem_alerta('Informe a medida no campo "Referência"')
                    self.line_Medida.setFocus()
                else:
                    self.item_produtomanual()
            elif embalagem == "SER":
                referencia = self.line_Referencia_Manu.text()
                if not referencia:
                    mensagem_alerta('O campo "Referência:" não pode estar vazio')
                else:
                    self.item_produtomanual()
            else:
                qtdezinha = self.line_Qtde_Manu.text()

                if len(qtdezinha) == 0:
                    mensagem_alerta('O campo "Qtde:" não pode estar vazio')
                    self.line_Qtde_Manu.clear()
                    self.line_Qtde_Manu.setFocus()
                elif qtdezinha == "0":
                    mensagem_alerta('O campo "Qtde:" não pode ser "0"')
                    self.line_Qtde_Manu.clear()
                    self.line_Qtde_Manu.setFocus()
                else:
                    self.item_produtomanual()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def item_produtomanual(self):
        try:
            unidadezinha = self.line_UM_Manu.text()
            codiguzinho = self.line_Codigo_Manu.text()
            codigo_produto = int(codiguzinho)
            descricaozinho = self.line_Descricao_Manu.text()

            dados_prod = self.tab_prod.consulta_por_codigo(codigo_produto)
            embalagem = dados_prod[0][7]
            um = dados_prod[0][4]
            kg_mt = dados_prod[0][8]

            if embalagem == "SIM":
                line_unidade = self.line_Unidade.text()
                combo_unidade = self.combo_Unidade.currentText()
                if line_unidade == "1":
                    um_texto = line_unidade + " " + combo_unidade.upper()
                else:
                    um_texto = line_unidade + " " + combo_unidade.upper() + "S"

                line_medida = self.line_Medida.text()
                combo_medida = self.combo_Medida.currentText()

                if line_medida == "1":
                    dois_texto = line_medida + " " + combo_medida
                else:
                    dois_texto = line_medida + " " + combo_medida + "S"

                texto_completo = um_texto + " DE " + dois_texto

                qtdezinha_float = 0
                if um == "KG":
                    if combo_medida == "MT":
                        mts_total = valores_para_float(line_unidade) * valores_para_float(line_medida)
                    else:
                        mts_total = valores_para_float(line_unidade) * (valores_para_float(line_medida) / 1000)

                    kg_total = mts_total * valores_para_float(kg_mt)
                    qtdezinha_float = "%.2f" % kg_total

                elif um == "MT":
                    if combo_medida == "MT":
                        mts_total = valores_para_float(line_unidade) * valores_para_float(line_medida)
                    else:
                        mts_total = valores_para_float(line_unidade) * (valores_para_float(line_medida) / 1000)

                    qtdezinha_float = "%.2f" % mts_total

                elif um == "MM":
                    if combo_medida == "MT":
                        mts_total = valores_para_float(line_unidade) * valores_para_float(line_medida)
                    else:
                        mts_total = valores_para_float(line_unidade) * (valores_para_float(line_medida) / 1000)

                    mm_total = mts_total * 1000

                    qtdezinha_float = "%.2f" % mm_total
                else:
                    mensagem_alerta('A unidade de medida do produto não está prevista!')

                referencia_certa = texto_completo

            elif embalagem == "SER":
                ref_servico = self.line_Referencia_Manu.text()
                referencia_certa = ref_servico.upper()

                qtdezinha = self.line_Qtde_Manu.text()
                qtdezinha_float = valores_para_float(qtdezinha)
            else:
                referencia_certa = self.line_Referencia_Manu.text()
                qtdezinha = self.line_Qtde_Manu.text()
                qtdezinha_float = valores_para_float(qtdezinha)

            destino = self.line_Destino_Manu.text()
            if len(destino) == 0:
                mensagem_alerta('O campo "Destino:" não pode estar vazio')
            else:
                destino_maiuscula = destino.upper()
                destino_certo = unidecode(destino_maiuscula)
                dados = [codiguzinho, descricaozinho, referencia_certa, unidadezinha, qtdezinha_float,
                         destino_certo]

                dados_prod = self.tab_prod.consulta_por_codigo(codigo_produto)
                ncm = dados_prod[0][9]
                if ncm:
                    existe_compra = self.consulta_compras_pendentes(codigo_produto)
                    if existe_compra:
                        extrai_solicitacao = extrair_tabela(self.table_Solicitacao)

                        ja_existe = False
                        for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                            if c_sol == codiguzinho:
                                ja_existe = True
                                break

                        if not ja_existe:
                            extrai_solicitacao.append(dados)

                            self.procedimento_lanca_tabela(extrai_solicitacao)
                        else:
                            mensagem_alerta(f'O item selecionado já está presente na tabela'
                                            f'"Lista Solicitação".')
                            extrai_solicitacao.append(dados)

                            self.procedimento_lanca_tabela(extrai_solicitacao)

                        self.line_Codigo_Manu.clear()
                        self.line_Descricao_Manu.clear()
                        self.line_Referencia_Manu.clear()
                        self.line_UM_Manu.clear()
                        self.line_Local_Manu.clear()
                        self.line_Qtde_Manu.clear()
                        self.line_Saldo_Manu.clear()
                        self.line_Referencia_Manu.clear()
                        self.line_Codigo_Manu.setFocus()
                        self.btn_Consome_Manu.setEnabled(False)
                        self.desaparece_referencia_editada()
                else:
                    mensagem_alerta(f'Este produto está sem "NCM" no cadastro.\n\n'
                                    f'Aproveite para verificar se o produto está apto '
                                    f'para compra no Siger.')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def limpa_tudo(self):
        try:
            self.table_Recomendacao.clearContents()
            self.line_Codigo_Manu.clear()
            self.line_Descricao_Manu.clear()
            self.line_Referencia_Manu.clear()
            self.line_UM_Manu.clear()
            self.line_Qtde_Manu.clear()
            self.line_Saldo_Manu.clear()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def procedimento_lanca_tabela(self, extrai_solicitacao):
        try:
            lanca_tabela(self.table_Solicitacao, extrai_solicitacao)
            self.pintar_bloquear_tabela_solicitacao()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def excluir_tudo_sol(self):
        try:
            extrai_sol = extrair_tabela(self.table_Solicitacao)
            if not extrai_sol:
                mensagem_alerta(f'A tabela "Lista Solicitação" está vazia!')
            else:
                self.table_Solicitacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def excluir_tudo_rec(self):
        try:
            extrai_recomendados = extrair_tabela(self.table_Recomendacao)
            if not extrai_recomendados:
                mensagem_alerta(f'A tabela "Lista de Recomendações" está vazia!')
            else:
                self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def excluir_item_sol(self):
        try:
            extrai_sol = extrair_tabela(self.table_Solicitacao)
            if not extrai_sol:
                mensagem_alerta(f'A tabela "Lista Solicitação" está vazia!')
            else:
                linha_selecao = self.table_Solicitacao.currentRow()
                if linha_selecao >= 0:
                    self.table_Solicitacao.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def excluir_item_rec(self):
        try:
            extrai_recomendados = extrair_tabela(self.table_Recomendacao)
            if not extrai_recomendados:
                mensagem_alerta(f'A tabela "Lista de Recomendações" está vazia!')
            else:
                linha_selecao = self.table_Recomendacao.currentRow()
                if linha_selecao >= 0:
                    self.table_Recomendacao.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def excluir_item_anexo(self):
        try:
            extrai_tabela_anexo = extrair_tabela(self.table_Anexos)
            if not extrai_tabela_anexo:
                mensagem_alerta(f'A tabela "Lista Anexos" está vazia!')
            else:
                linha_selecao = self.table_Anexos.currentRow()
                if linha_selecao >= 0:
                    self.table_Anexos.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def tudo_verifica_deondevem(self):
        try:
            sender = self.sender()

            if sender == self.btn_Adicionar_Todos:
                extrai_recomendados = extrair_tabela(self.table_Recomendacao)
                if not extrai_recomendados:
                    mensagem_alerta(f'A tabela "Lista de Recomendações" está vazia!')

                else:
                    combo_consumo = self.combo_Consumo.currentText()
                    anos = self.line_Ano_Consumo.text()

                    desc_pai = self.line_Descricao_Estrut.text()
                    qtde_estrut = self.line_Qtde_Estrut.text()
                    if desc_pai and qtde_estrut:
                        self.lancar_tudo_estrutura()
                    elif combo_consumo and anos:
                        self.lancar_tudo_consumo_interno()
                    elif not desc_pai or not qtde_estrut:
                        mensagem_alerta(f'Os campos "Código" e "Qtde" da estrutura precisam'
                                        f'estar preenchidos!')
                    elif not combo_consumo or not anos:
                        mensagem_alerta(f'Os campos "Setor" e "Compra desde" do Consumo Interno '
                                        f'precisam estar preenchidos!')

            elif sender == self.btn_Adicionar_SemSaldo:
                extrai_recomendados = extrair_tabela(self.table_Recomendacao)
                if not extrai_recomendados:
                    mensagem_alerta(f'A tabela "Lista de Recomendações" está vazia!')

                else:
                    combo_consumo = self.combo_Consumo.currentText()
                    anos = self.line_Ano_Consumo.text()

                    desc_pai = self.line_Descricao_Estrut.text()
                    qtde_estrut = self.line_Qtde_Estrut.text()
                    if desc_pai and qtde_estrut:
                        self.lancar_semsaldo_estrutura()
                    elif combo_consumo and anos:
                        self.lancar_semsaldo_consumo_interno()
                    elif not desc_pai or not qtde_estrut:
                        mensagem_alerta(f'Os campos "Código" e "Qtde" da estrutura precisam'
                                        f'estar preenchidos!')
                    elif not combo_consumo or not anos:
                        mensagem_alerta(f'Os campos "Setor" e "Compra desde" do Consumo Interno '
                                        f'precisam estar preenchidos!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lanca_so_chapas(self):
        try:
            nova_tabela = []

            extrai_recomendados = extrair_tabela(self.table_Recomendacao)

            if extrai_recomendados:
                for dados in extrai_recomendados:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    dados_prod = self.tab_prod.consulta_por_codigo(cod)
                    if dados_prod:
                        tipo = dados_prod[0][12]

                        if tipo == 116:
                            didis = (cod, desc, ref, um, qtde, local, saldo)
                            nova_tabela.append(didis)
                        elif tipo == 84:
                            didis = (cod, desc, ref, um, qtde, local, saldo)
                            nova_tabela.append(didis)
                        elif tipo == 85:
                            didis = (cod, desc, ref, um, qtde, local, saldo)
                            nova_tabela.append(didis)

            ja_foi = 0
            itens_foi = []
            if nova_tabela:
                desc_pai = self.line_Descricao_Estrut.text()

                for dados in nova_tabela:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    dados_prod = self.tab_prod.consulta_por_codigo(cod)
                    ncm = dados_prod[0][9]
                    if ncm:
                        existe_compra = self.consulta_compras_pendentes(cod)
                        if existe_compra:
                            ja_existe = False

                            extrai_solicitacao = extrair_tabela(self.table_Solicitacao)
                            if extrai_solicitacao:
                                for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                                    if c_sol == cod:
                                        ja_existe = True
                                        break

                            if not ja_existe:
                                didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                extrai_solicitacao.append(didis)

                            else:
                                ja_foi = ja_foi + 1
                                didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                itens_foi.append(didis)
                                extrai_solicitacao.append(didis)

                            self.procedimento_lanca_tabela(extrai_solicitacao)
                        else:
                            break
                    else:
                        mensagem_alerta(f'O produto {cod} está sem "NCM" no cadastro.\n\n'
                                        f'Aproveite para verificar se o produto está apto '
                                        f'para compra no Siger.')
                        break

            if ja_foi > 0:
                mensagem = ""
                if len(itens_foi) == 1:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod

                    mensagem_alerta(f'O item {mensagem} já está presente na tabela'
                                    f'"Lista Solicitação".')
                else:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod + ", "

                    mensagem_alerta(f'Os itens {mensagem} já estão presentes na tabela'
                                    f'"Lista Solicitação".')
            self.combo_Consumo.setCurrentText("")
            self.reiniciando_produto_estrutura()
            self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lancar_tudo_estrutura(self):
        try:
            ja_foi = 0
            itens_foi = []

            extrai_recomendados = extrair_tabela(self.table_Recomendacao)

            if extrai_recomendados:
                desc_pai = self.line_Descricao_Estrut.text()

                for dados in extrai_recomendados:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    dados_prod = self.tab_prod.consulta_por_codigo(cod)
                    ncm = dados_prod[0][9]
                    if ncm:
                        existe_compra = self.consulta_compras_pendentes(cod)
                        if existe_compra:
                            ja_existe = False

                            extrai_solicitacao = extrair_tabela(self.table_Solicitacao)
                            if extrai_solicitacao:
                                for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                                    if c_sol == cod:
                                        ja_existe = True
                                        break

                            if not ja_existe:
                                didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                extrai_solicitacao.append(didis)

                            else:
                                ja_foi = ja_foi + 1
                                didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                itens_foi.append(didis)
                                extrai_solicitacao.append(didis)

                            self.procedimento_lanca_tabela(extrai_solicitacao)

                        else:
                            break
                    else:
                        mensagem_alerta(f'O produto {cod} está sem "NCM" no cadastro.\n\n'
                                        f'Aproveite para verificar se o produto está apto '
                                        f'para compra no Siger.')
                        break

            if ja_foi > 0:
                mensagem = ""
                if len(itens_foi) == 1:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod

                    mensagem_alerta(f'O item {mensagem} já está presente na tabela'
                                    f'"Lista Solicitação".')
                else:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod + ", "

                    mensagem_alerta(f'Os itens {mensagem} já estão presentes na tabela'
                                    f'"Lista Solicitação".')
            self.combo_Consumo.setCurrentText("")
            self.reiniciando_produto_estrutura()
            self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lancar_tudo_consumo_interno(self):
        try:
            ja_foi = 0
            itens_foi = []

            extrai_recomendados = extrair_tabela(self.table_Recomendacao)

            if extrai_recomendados:
                origem = self.combo_Consumo.currentText()

                origemtete = origem.find(" - ")
                nome_origem = origem[origemtete:]

                destino_final = "CI" + nome_origem

                for dados in extrai_recomendados:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    dados_prod = self.tab_prod.consulta_por_codigo(cod)
                    ncm = dados_prod[0][9]
                    if ncm:
                        existe_compra = self.consulta_compras_pendentes(cod)
                        if existe_compra:
                            ja_existe = False

                            extrai_solicitacao = extrair_tabela(self.table_Solicitacao)
                            if extrai_solicitacao:
                                for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                                    if c_sol == cod:
                                        ja_existe = True
                                        break

                            if not ja_existe:
                                didis = [cod, desc, ref, um, float(qtde), destino_final]
                                extrai_solicitacao.append(didis)
                            else:
                                ja_foi = ja_foi + 1
                                didis = [cod, desc, ref, um, float(qtde), destino_final]
                                itens_foi.append(didis)
                                extrai_solicitacao.append(didis)

                            self.procedimento_lanca_tabela(extrai_solicitacao)

                        else:
                            break
                    else:
                        mensagem_alerta(f'O produto {cod} está sem "NCM" no cadastro.\n\n'
                                        f'Aproveite para verificar se o produto está apto '
                                        f'para compra no Siger.')
                        break

            if ja_foi > 0:
                mensagem = ""
                if len(itens_foi) == 1:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod

                    mensagem_alerta(f'O item {mensagem} já está presente na tabela'
                                    f'"Lista Solicitação".')
                else:
                    for didi in itens_foi:
                        cod, desc, ref, um, qtde, destino = didi

                        mensagem = mensagem + cod + ", "

                    mensagem_alerta(f'Os itens {mensagem} já estão presentes na tabela'
                                    f'"Lista Solicitação".')
            self.combo_Consumo.setCurrentText("")
            self.reiniciando_produto_estrutura()
            self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lancar_semsaldo_estrutura(self):
        try:
            ja_foi = 0
            sem_ncm = 0
            itens_foi = []
            ja_existe = False

            extrai_recomendados = extrair_tabela(self.table_Recomendacao)
            extrai_solicitacao = extrair_tabela(self.table_Solicitacao)
            qtde_extrai_velho = len(extrai_solicitacao)

            if extrai_recomendados:
                desc_pai = self.line_Descricao_Estrut.text()

                for dados in extrai_recomendados:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    if float(saldo) < float(qtde):
                        dados_prod = self.tab_prod.consulta_por_codigo(cod)
                        ncm = dados_prod[0][9]
                        if ncm:
                            existe_compra = self.consulta_compras_pendentes(cod)
                            if existe_compra:
                                if extrai_solicitacao:
                                    for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                                        if c_sol == cod:
                                            ja_existe = True
                                            break

                                if not ja_existe:
                                    didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                    extrai_solicitacao.append(didis)
                                else:
                                    ja_foi = ja_foi + 1
                                    didis = [cod, desc, ref, um, float(qtde), desc_pai]
                                    itens_foi.append(didis)
                                    extrai_solicitacao.append(didis)
                            else:
                                sem_ncm = sem_ncm + 1
                                break
                        else:
                            sem_ncm = sem_ncm + 1
                            mensagem_alerta(f'O produto {cod} está sem "NCM" no cadastro.\n\n'
                                            f'Aproveite para verificar se o produto está apto '
                                            f'para compra no Siger.')
                            break

            if sem_ncm == 0:
                qtde_extrai_novo = len(extrai_solicitacao)

                if qtde_extrai_velho == qtde_extrai_novo:
                    mensagem_alerta(f'Não existem itens sem saldo".')
                else:
                    self.procedimento_lanca_tabela(extrai_solicitacao)

                if ja_foi > 0:
                    mensagem = ""
                    if len(itens_foi) == 1:
                        for didi in itens_foi:
                            cod, desc, ref, um, qtde, destino = didi

                            mensagem = mensagem + cod

                        mensagem_alerta(f'O item {mensagem} já está presente na tabela'
                                        f'"Lista Solicitação".')
                    else:
                        for didi in itens_foi:
                            cod, desc, ref, um, qtde, destino = didi

                            mensagem = mensagem + cod + ", "

                        mensagem_alerta(f'Os itens {mensagem} já estão presentes na tabela'
                                        f'"Lista Solicitação".')
                self.combo_Consumo.setCurrentText("")
                self.reiniciando_produto_estrutura()
                self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def lancar_semsaldo_consumo_interno(self):
        try:
            ja_foi = 0
            sem_ncm = 0
            itens_foi = []
            ja_existe = False

            extrai_recomendados = extrair_tabela(self.table_Recomendacao)
            extrai_solicitacao = extrair_tabela(self.table_Solicitacao)
            qtde_extrai_velho = len(extrai_solicitacao)

            if extrai_recomendados:
                origem = self.combo_Consumo.currentText()
                origemtete = origem.find(" - ")
                nome_origem = origem[origemtete:]

                destino_final = "CI" + nome_origem

                for dados in extrai_recomendados:
                    cod, desc, ref, um, qtde, local, saldo = dados

                    if float(saldo) < float(qtde):
                        dados_prod = self.tab_prod.consulta_por_codigo(cod)
                        ncm = dados_prod[0][9]
                        if ncm:
                            existe_compra = self.consulta_compras_pendentes(cod)
                            if existe_compra:
                                if extrai_solicitacao:
                                    for c_sol, d_sol, ref_sol, um_sol, qtde_sol, de_sol in extrai_solicitacao:
                                        if c_sol == cod:
                                            ja_existe = True
                                            break

                                if not ja_existe:
                                    didis = [cod, desc, ref, um, float(qtde), destino_final]
                                    extrai_solicitacao.append(didis)
                                else:
                                    ja_foi = ja_foi + 1
                                    didis = [cod, desc, ref, um, float(qtde), destino_final]
                                    itens_foi.append(didis)
                                    extrai_solicitacao.append(didis)

                            else:
                                break
                        else:
                            sem_ncm = sem_ncm + 1
                            mensagem_alerta(f'O produto {cod} está sem "NCM" no cadastro.\n\n'
                                            f'Aproveite para verificar se o produto está apto '
                                            f'para compra no Siger.')
                            break

            if sem_ncm == 0:
                qtde_extrai_novo = len(extrai_solicitacao)

                if qtde_extrai_velho == qtde_extrai_novo:
                    mensagem_alerta(f'Não existem itens sem saldo".')
                else:
                    self.procedimento_lanca_tabela(extrai_solicitacao)

                if ja_foi > 0:
                    mensagem = ""
                    if len(itens_foi) == 1:
                        for didi in itens_foi:
                            cod, desc, ref, um, qtde, destino = didi

                            mensagem = mensagem + cod

                        mensagem_alerta(f'O item {mensagem} já está presente na tabela'
                                        f'"Lista Solicitação".')
                    else:
                        for didi in itens_foi:
                            cod, desc, ref, um, qtde, destino = didi

                            mensagem = mensagem + cod + ", "

                        mensagem_alerta(f'Os itens {mensagem} já estão presentes na tabela'
                                        f'"Lista Solicitação".')
                self.combo_Consumo.setCurrentText("")
                self.reiniciando_produto_estrutura()
                self.table_Recomendacao.setRowCount(0)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def verifica_salvamento(self):
        try:
            extrai_sol = extrair_tabela(self.table_Solicitacao)
            if not extrai_sol:
                mensagem_alerta(f'A tabela "Lista Solicitação" está vazia!')
            else:
                num_sol = self.line_Num_Sol.text()

                testar_erros = 0

                if not num_sol:
                    testar_erros = testar_erros + 1
                    mensagem_alerta(f'O campo "Nº SOL:" não pode estar vazio!')
                elif num_sol == "0":
                    testar_erros = testar_erros + 1
                    mensagem_alerta(f'O "Nº SOL:" não pode ser "0"!')

                for itens in extrai_sol:
                    codigo, descricao, referencia, um, qtde, destino = itens

                    if not qtde:
                        mensagem_alerta('Na Tabela "Lista Solicitação" possui '
                                        'produtos sem quantidade!')
                        testar_erros = testar_erros + 1
                        break
                    else:
                        dados_prod = self.tab_prod.consulta_por_codigo(codigo)

                        if dados_prod:
                            embalagem = dados_prod[0][7]

                            if embalagem == "SIM":
                                if not referencia:
                                    mensagem_alerta('O campo "Referência" pintado de amarelo, '
                                                    'não pode estar vazio!')
                                    testar_erros = testar_erros + 1
                                    break
                            elif embalagem == "SER":
                                if not referencia:
                                    mensagem_alerta('O campo "Referência" pintado de amarelo, '
                                                    'não pode estar vazio!')
                                    testar_erros = testar_erros + 1
                                    break

                if testar_erros == 0:
                    self.salvar_lista()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def salvar_lista(self):
        try:
            observacao_requisicao = self.line_Obs.text()
            obs_req_maiuscula = observacao_requisicao.upper()
            observacao_certa = unidecode(obs_req_maiuscula)
            nome_computador = socket.gethostname()

            datamov = self.date_Emissao.text()
            date_mov = datetime.strptime(datamov, '%d/%m/%Y').date()
            data_mov_certa = str(date_mov)
            data_mov_certa2 = "'" + data_mov_certa + "'"

            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_ORDEMSOLICITACAO_ID,0) from rdb$database;")
            ultimo_req0 = cursor.fetchall()
            ultimo_req1 = ultimo_req0[0]
            ultimo_req = int(ultimo_req1[0]) + 1

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ordemsolicitacao (IDSOLICITACAO, DATAEMISSAO, STATUS, OBS, NOME_PC) "
                           f"values (GEN_ID(GEN_ORDEMSOLICITACAO_ID,1), "
                           f"{data_mov_certa2}, 'A', '{observacao_certa}', '{nome_computador}');")

            dados_alterados = extrair_tabela(self.table_Solicitacao)

            for itens in dados_alterados:
                codigo, descricao, referencia, um, qtde, destino = itens

                descricao_maiuscula = descricao.upper()
                descricao_certa = unidecode(descricao_maiuscula)

                referencia_maiuscula = referencia.upper()
                referencia_certa = unidecode(referencia_maiuscula)

                um_maiuscula = um.upper()
                um_certa = unidecode(um_maiuscula)

                destino_maiuscula = destino.upper()
                destino_certa = unidecode(destino_maiuscula)

                if "," in qtde:
                    qtdezinha_com_ponto = qtde.replace(',', '.')
                    qtdezinha_float = float(qtdezinha_com_ponto)
                else:
                    qtdezinha_float = float(qtde)

                dados_prod = self.tab_prod.consulta_por_codigo(codigo)
                id_prod = dados_prod[0][0]
                embalagem = dados_prod[0][7]

                if codigo == "1":
                    cursor = conecta.cursor()
                    cursor.execute(f"Insert into produtoordemsolicitacao (ID, MESTRE, PRODUTO, DESCRICAO, REFERENCIA, "
                                   f"UM, QUANTIDADE, DATA, STATUS, DESTINO) "
                                   f"values (GEN_ID(GEN_PRODUTOORDEMSOLICITACAO_ID,1), {ultimo_req}, "
                                   f"{id_prod}, '{descricao_certa}', '{referencia_certa}', '{um_certa}', "
                                   f"{qtdezinha_float}, {data_mov_certa2}, 'A', '{destino_certa}');")

                elif embalagem == "SIM":
                    cursor = conecta.cursor()
                    cursor.execute(f"Insert into produtoordemsolicitacao (ID, MESTRE, PRODUTO, REFERENCIA, "
                                   f"QUANTIDADE, DATA, STATUS, DESTINO) "
                                   f"values (GEN_ID(GEN_PRODUTOORDEMSOLICITACAO_ID,1), {ultimo_req}, "
                                   f"{id_prod}, '{referencia_certa}', {qtdezinha_float}, "
                                   f"{data_mov_certa2}, 'A', '{destino_certa}');")
                else:
                    cursor = conecta.cursor()
                    cursor.execute(f"Insert into produtoordemsolicitacao (ID, MESTRE, PRODUTO, QUANTIDADE, "
                                   f"DATA, STATUS, DESTINO) "
                                   f"values (GEN_ID(GEN_PRODUTOORDEMSOLICITACAO_ID,1), {ultimo_req}, "
                                   f"{id_prod}, {qtdezinha_float}, {data_mov_certa2}, "
                                   f"'A', '{destino_certa}');")

            self.grava_anexo(ultimo_req)
            conecta.commit()

            self.reiniciando_sol()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def reiniciando_sol(self):
        try:
            self.table_Recomendacao.setRowCount(0)
            self.table_Solicitacao.setRowCount(0)
            self.table_Anexos.setRowCount(0)

            self.line_Obs.clear()

            self.line_Descricao_Manu.clear()
            self.line_Referencia_Manu.clear()
            self.line_Saldo_Manu.clear()
            self.line_UM_Manu.clear()
            self.line_Codigo_Manu.clear()
            self.line_Local_Manu.clear()
            self.line_Destino_Manu.clear()
            self.line_Qtde_Manu.clear()
            self.line_Referencia_Manu.clear()

            data_hoje = date.today()
            ano_atual = data_hoje.strftime("%Y")
            ano_menos_dois = str(int(ano_atual) - 2)
            self.line_Ano_Consumo.setText(ano_menos_dois)

            self.line_Codigo_Manu.setFocus()

            self.definir_validador_lineedit()
            self.definir_lineedit_bloqueados()
            self.definir_line_ano_consumo()
            definir_proximo_generator(self.line_Num_Sol, "ORDEMSOLICITACAO")
            definir_data_atual(self.date_Emissao)
            self.definir_combo_consumo()
            self.desaparece_referencia_editada()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def grava_anexo(self, num_sol):
        try:
            lista_anexos = extrair_tabela(self.table_Anexos)
            if lista_anexos:
                for anexo in lista_anexos:
                    nome_arquivo, caminho = anexo

                    nome_arquivo_usuario = os.path.basename(caminho)
                    nome_arquivo_final = f'{num_sol} - {nome_arquivo_usuario}'

                    destination_path = os.path.join(r'\\PUBLICO\Python\0 - Versões Antigas\anexos', nome_arquivo_final)

                    shutil.copy2(caminho, destination_path)

                    cursor = conecta.cursor()
                    cursor.execute(f"Insert into SOLICITACAO_ANEXO (ID, CAMINHO, ID_SOLICITACAO) "
                                   f"values (GEN_ID(GEN_SOLICITACAO_ANEXO_ID,1), '{destination_path}', {num_sol});")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)

    def gera_excel(self):
        try:
            num_sol = self.line_Num_Sol.text()
            obs_solicitacao = self.line_Obs.text()

            data_hoje = date.today()
            data_certa = data_hoje.strftime("%d/%m/%Y")

            if not obs_solicitacao:
                obs_sol = ""
            else:
                obs_sol_maiuscula = obs_solicitacao.upper()
                obs_sol = unidecode(obs_sol_maiuscula)

            dados_tabela = extrair_tabela(self.table_Solicitacao)
            d_um = []

            embalagem_sim_rows = []

            for tabi in dados_tabela:
                cod_1, desc_1, ref_1, um_1, qtde_1, destino = tabi
                if "," in qtde_1:
                    qtdezinha_com_ponto = qtde_1.replace(',', '.')
                    qtdezinha_float = float(qtdezinha_com_ponto)
                else:
                    qtdezinha_float = float(qtde_1)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, codigo, embalagem FROM produto where codigo = '{cod_1}';")
                dados_produto = cursor.fetchall()
                if dados_produto:
                    id_produto, codigo, embalagem = dados_produto[0]
                    if embalagem == "SIM" or embalagem == "SER":
                        embalagem_sim_rows.append(len(d_um) - 1)

                dados = (cod_1, desc_1, ref_1, um_1, qtdezinha_float, destino)
                d_um.append(dados)

            df = pd.DataFrame(d_um, columns=['Código', 'Descrição', 'Referência', 'UM', 'Qtde', 'Destino'])

            codigo_int = {'Código': int}
            df = df.astype(codigo_int)
            qtde_float = {'Qtde': float}
            df = df.astype(qtde_float)

            book = load_workbook('Modelo Solicitação.xlsx')

            desktop = Path.home() / "Desktop"
            desk_str = str(desktop)
            nome_req = f'\Solicitação {num_sol}.xlsx'
            caminho = (desk_str + nome_req)

            writer = pd.ExcelWriter(caminho, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            linhas_frame = df.shape[0]
            colunas_frame = df.shape[1]

            linhas_certas = linhas_frame + 2 + 9
            colunas_certas = colunas_frame + 1

            ws = book.active

            inicia = 11
            rows = range(inicia, inicia + linhas_frame)
            columns = range(1, colunas_certas)

            ws.row_dimensions[linhas_certas + 2].height = 30
            ws.row_dimensions[linhas_certas + 4].height = 30

            for row in rows:
                for col in columns:
                    ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
                    ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                      right=Side(border_style='thin', color='00000000'),
                                                      top=Side(border_style='thin', color='00000000'),
                                                      bottom=Side(border_style='thin', color='00000000'),
                                                      diagonal=Side(border_style='thick', color='00000000'),
                                                      diagonal_direction=0,
                                                      outline=Side(border_style='thin', color='00000000'),
                                                      vertical=Side(border_style='thin', color='00000000'),
                                                      horizontal=Side(border_style='thin', color='00000000'))

            ws.merge_cells(f'A8:D8')
            top_left_cell = ws[f'A8']
            c = ws[f'A8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Solicitação Nº  ' + num_sol

            ws.merge_cells(f'E8:F8')
            top_left_cell = ws[f'E8']
            c = ws[f'E8']
            c.alignment = Alignment(horizontal='center',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=14, bold=True)
            top_left_cell.value = 'Emissão:  ' + data_certa

            ws.merge_cells(f'B{linhas_certas + 2}:B{linhas_certas + 2}')
            top_left_cell = ws[f'B{linhas_certas + 2}']
            c = ws[f'B{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='right',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=True)
            top_left_cell.value = "Observação:  "

            ws.merge_cells(f'C{linhas_certas + 2}:H{linhas_certas + 2}')
            top_left_cell = ws[f'C{linhas_certas + 2}']
            c = ws[f'C{linhas_certas + 2}']
            c.alignment = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=False,
                                    shrink_to_fit=False,
                                    indent=0)
            c.font = Font(size=12, bold=False)
            top_left_cell.value = obs_sol

            df.to_excel(writer, 'Sheet1', startrow=10, startcol=0, header=False, index=False)

            for row_idx in embalagem_sim_rows:
                row = row_idx + 12
                col = 3
                ws.cell(row, col).fill = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            writer.save()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            tratar_notificar_erros(e, nome_funcao, self.nome_arquivo)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaSolIncluir()
    tela.show()
    qt.exec_()
