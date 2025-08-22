import sys
from banco_dados.conexao import conecta
from forms.tela_ci_consumo_temporario_op import *
from banco_dados.controle_erros import grava_erro_banco
from comandos.tabelas import lanca_tabela, layout_cabec_tab, extrair_tabela
from comandos.telas import tamanho_aplicacao, icone
from comandos.lines import validador_so_numeros
from comandos.conversores import valores_para_float
from arquivos.chamar_arquivos import definir_caminho_arquivo
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
import inspect
import os
from datetime import date, datetime
import traceback
import socket
import getpass


class TelaConsumoTemporario(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        self.nome_computador = socket.gethostname()
        self.username = getpass.getuser()

        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_producao.png")
        tamanho_aplicacao(self)
        layout_cabec_tab(self.table_Lista)

        self.processando = False

        self.definir_data_emissao()

        self.table_Lista.viewport().installEventFilter(self)

        self.btn_Adicionar.clicked.connect(self.lanca_produto_na_tabela)

        self.btn_Salvar.clicked.connect(self.verifica_salvamento)
        self.btn_Limpar.clicked.connect(self.reiniciando_tela)
        self.btn_Excluir.clicked.connect(self.excluir_item)

        self.btn_Excel.clicked.connect(self.gera_excel)

        self.definir_combo_funcionario()

        validador_so_numeros(self.line_Num)

        self.line_Num.returnPressed.connect(lambda: self.verifica_op())
        self.line_Codigo.editingFinished.connect(self.verifica_line_codigo)
        self.line_Qtde.editingFinished.connect(self.verifica_line_qtde_manual)

        self.combo_Funcionario.activated.connect(self.lanca_produto_na_tabela)

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def pergunta_confirmacao(self, mensagem):
        try:
            confirmacao = QMessageBox()
            confirmacao.setIcon(QMessageBox.Question)
            confirmacao.setText(mensagem)
            confirmacao.setWindowTitle("Confirmação")

            sim_button = confirmacao.addButton("Sim", QMessageBox.YesRole)
            nao_button = confirmacao.addButton("Não", QMessageBox.NoRole)

            confirmacao.setDefaultButton(nao_button)

            confirmacao.exec_()

            if confirmacao.clickedButton() == sim_button:
                return True
            else:
                return False

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_data_emissao(self):
        try:
            data_hoje = date.today()
            self.date_Emissao.setDate(data_hoje)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_combo_funcionario(self):
        try:
            tabela = []

            self.combo_Funcionario.clear()
            tabela.append("")

            cur = conecta.cursor()
            cur.execute(f"SELECT id, funcionario FROM funcionarios where ativo = 'S' order by funcionario;")
            detalhes_func = cur.fetchall()

            for dadus in detalhes_func:
                ides, func = dadus
                tabela.append(func)

            self.combo_Funcionario.addItems(tabela)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_op(self):
        try:
            self.limpa_dados_produto()
            self.label_6.setText('')

            numero_os_line = self.line_Num.text()
            if len(numero_os_line) == 0:
                self.mensagem_alerta('O campo "Nº OP" não pode estar vazio')
                self.reiniciando_tela()
            elif int(numero_os_line) == 0:
                self.mensagem_alerta('O campo "Nº OP" não pode ser "0"')
                self.reiniciando_tela()
            else:
                self.verifica_sql_op()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_sql_op(self):
        try:
            numero_op = self.line_Num.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT numero, datainicial, status, produto, quantidade "
                           f"FROM ordemservico where numero = {numero_op};")
            extrair_dados = cursor.fetchall()
            if not extrair_dados:
                self.mensagem_alerta('Este número de "OP" não existe!')
                self.reiniciando_tela()
            else:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT numero, datainicial, status, produto, quantidade "
                               f"FROM ordemservico where numero = {numero_op} AND status = 'A';")
                select_status = cursor.fetchall()

                if not select_status:
                    self.mensagem_alerta('Esta Ordem de Produção está encerrada!')
                    self.reiniciando_tela()
                else:
                    self.lanca_dados_op()
                    self.line_Codigo.setFocus()

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT COALESCE((extract(day from aux.data)||'/'||"
                                   f"extract(month from aux.data)||'/'||"
                                   f"extract(year from aux.data)), '') AS DATA, prod.codigo, prod.descricao, "
                                   f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                                   f"aux.qtde, func.funcionario "
                                   f"FROM CONSUMO_TEMPORARIO_OP as aux "
                                   f"INNER JOIN produto prod ON aux.id_produto = prod.id "
                                   f"INNER JOIN FUNCIONARIOS func ON aux.id_funcionario = func.id "
                                   f"where aux.num_op = {numero_op};")
                    dados_aux = cursor.fetchall()

                    if dados_aux:
                        lanca_tabela(self.table_Lista, dados_aux)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def dados_op(self):
        try:
            numero_os_line = self.line_Num.text()
            cur = conecta.cursor()
            cur.execute(f"SELECT id, numero, datainicial, status, produto, quantidade, id_estrutura "
                        f"FROM ordemservico where numero = {numero_os_line};")
            extrair_dados = cur.fetchall()
            id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, id_estrut = extrair_dados[0]

            return id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, id_estrut

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_op(self):
        try:
            id_os, numero_os, data_emissao, status_os, produto_os, qtde_os, id_estrut = self.dados_op()

            cur = conecta.cursor()
            cur.execute(f"SELECT codigo, descricao, COALESCE(obs, ' ') as obs, "
                        f"COALESCE(descricaocomplementar, ' ') as compl, unidade, id_versao "
                        f"FROM produto where id = {produto_os};")
            detalhes_produtos = cur.fetchall()

            codigo_id, descricao_id, referencia_id, complementar_id, unidade_id, id_versao = detalhes_produtos[0]
            self.line_Codigo_3.setText(codigo_id)
            self.line_Descricao_3.setText(descricao_id)
            self.line_Referencia_3.setText(referencia_id)
            self.line_UM_3.setText(unidade_id)
            numero = str(qtde_os).replace('.', ',')
            self.line_Qtde_3.setText(numero)

            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, num_versao, data_versao, obs, data_criacao "
                           f"from estrutura "
                           f"where id = {id_versao};")
            tabela_versoes = cursor.fetchall()

            id_estrut, num_versao, data, obs, criacao = tabela_versoes[0]

            data_versao = data.strftime("%d/%m/%Y")

            if id_versao == id_estrut:
                status_txt = "ATIVO"
            else:
                status_txt = "INATIVO"

            msg = f"VERSÃO: {num_versao} - {data_versao} - {status_txt}"
            self.line_Versao_2.setText(msg)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_codigo(self):
        if not self.processando:
            try:
                self.processando = True

                self.label_6.setText('')

                codigo_produto = self.line_Codigo.text()

                if codigo_produto:
                    if int(codigo_produto) == 0:
                        self.mensagem_alerta('O campo "Código" não pode ser "0"')
                        self.limpa_dados_produto()
                    else:
                        self.verifica_sql_produto()

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False

    def verifica_sql_produto(self):
        try:
            codigo_produto = self.line_Codigo.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, descricao, COALESCE(obs, ' ') as obs, unidade, localizacao, quantidade "
                           f"FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cursor.fetchall()
            if not detalhes_produto:
                self.mensagem_alerta('Este código de produto não existe!')
                self.limpa_dados_produto()
            else:
                ides, descr, ref, um, local, saldo = detalhes_produto[0]
                saldo_float = valores_para_float(saldo)
                if saldo_float > 0:
                    self.lanca_dados_produto()
                else:
                    self.mensagem_alerta('Este produto não possui saldo em estoque!')
                    self.limpa_dados_produto()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_produto(self):
        try:
            codigo_produto = self.line_Codigo.text()

            cur = conecta.cursor()
            cur.execute(f"SELECT id, descricao, COALESCE(obs, '') as obs, unidade, COALESCE(localizacao, '') as local, "
                        f"quantidade, embalagem, kilosmetro FROM produto where codigo = {codigo_produto};")
            detalhes_produto = cur.fetchall()
            ides, descr, ref, um, local, qtde, embalagem, kg_mt = detalhes_produto[0]

            self.line_Descricao.setText(descr)
            self.line_UM.setText(um)
            self.line_Referencia.setText(ref)
            self.line_Qtde.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_qtde_manual(self):
        if not self.processando:
            try:
                self.processando = True

                qtdezinha = self.line_Qtde.text()

                if qtdezinha:
                    if qtdezinha == "0":
                        self.mensagem_alerta('O campo "Qtde:" não pode ser "0"')
                        self.line_Qtde.clear()
                        self.line_Qtde.setFocus()
                    else:
                        self.combo_Funcionario.setFocus()

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False

    def lanca_produto_na_tabela(self):
        try:
            self.label_6.setText('')

            entrega = self.date_Emissao.text()
            entrega_mov = datetime.strptime(entrega, '%d/%m/%Y').date()
            data_string = '{}/{}/{}'.format(entrega_mov.day, entrega_mov.month, entrega_mov.year)

            funcionario = self.combo_Funcionario.currentText()

            um = self.line_UM.text()
            cod = self.line_Codigo.text()
            descr = self.line_Descricao.text()
            ref = self.line_Referencia.text()

            qtde_manu = self.line_Qtde.text()

            if cod and qtde_manu and funcionario:
                if "," in qtde_manu:
                    qtde_manu_com_ponto = qtde_manu.replace(',', '.')
                    qtde_float = float(qtde_manu_com_ponto)
                else:
                    qtde_float = float(qtde_manu)

                extrai_consumo = extrair_tabela(self.table_Lista)

                codigos_existentes = [linha[1] for linha in extrai_consumo]
                if cod in codigos_existentes:
                    self.mensagem_alerta(f"O código {cod} já está lançado na tabela!")
                else:
                    dados = [data_string, cod, descr, ref, um, qtde_float, funcionario]
                    extrai_consumo.append(dados)
                    lanca_tabela(self.table_Lista, extrai_consumo)

                self.limpa_dados_produto()
                self.line_Codigo.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_dados_op(self):
        try:
            self.line_Codigo_3.clear()
            self.line_Referencia_3.clear()
            self.line_Descricao_3.clear()
            self.line_UM_3.clear()
            self.line_Qtde_3.clear()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_dados_produto(self):
        try:
            self.line_Codigo.clear()
            self.line_Referencia.clear()
            self.line_Descricao.clear()
            self.line_UM.clear()
            self.line_Qtde.clear()

            self.combo_Funcionario.setCurrentText("")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item(self):
        try:
            self.label_6.setText('')

            extrai_recomendados = extrair_tabela(self.table_Lista)
            if extrai_recomendados:
                linha_selecao = self.table_Lista.currentRow()
                if linha_selecao >= 0:
                    self.table_Lista.removeRow(linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def reiniciando_tela(self):
        try:
            self.line_Num.clear()

            self.limpa_dados_op()
            self.limpa_dados_produto()

            self.line_Num.setFocus()

            extrai_recomendados = extrair_tabela(self.table_Lista)
            if extrai_recomendados:
                self.table_Lista.setRowCount(0)

            self.label_6.setText('')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_salvamento(self):
        try:
            numero_op = self.line_Num.text()
            print(numero_op)

            if not numero_op:
                self.mensagem_alerta('O campo "Nº OP" não pode estar vazio!')
                self.line_Num.setFocus()
            elif numero_op == "0":
                self.mensagem_alerta('O campo "Nº OP" não pode ser "0"!')
                self.line_Num.clear()
                self.line_Num.setFocus()
            else:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT numero, datainicial, status, produto, quantidade "
                               f"FROM ordemservico where numero = {numero_op};")
                extrair_dados = cursor.fetchall()
                if not extrair_dados:
                    self.mensagem_alerta('Este número de "OP" não existe!')
                    self.reiniciando_tela()
                else:
                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT numero, datainicial, status, produto, quantidade "
                                   f"FROM ordemservico where numero = {numero_op} AND status = 'A';")
                    select_status = cursor.fetchall()

                    if not select_status:
                        self.mensagem_alerta('Esta Ordem de Produção está encerrada!')
                        self.reiniciando_tela()
                    else:
                        print(numero_op)
                        self.salvar_dados()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_dados(self):
        try:
            numero_op = self.line_Num.text()

            data = self.date_Emissao.text()
            data_consumo = datetime.strptime(data, '%d/%m/%Y').date()
            data_consumo_certo = str(data_consumo)

            dados_consumo = extrair_tabela(self.table_Lista)

            cursor = conecta.cursor()
            cursor.execute(f"SELECT aux.data, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, prod.unidade, "
                           f"aux.qtde, func.funcionario "
                           f"FROM CONSUMO_TEMPORARIO_OP as aux "
                           f"INNER JOIN produto prod ON aux.id_produto = prod.id "
                           f"INNER JOIN FUNCIONARIOS func ON aux.id_funcionario = func.id "
                           f"where aux.num_op = {numero_op};")
            dados_aux = cursor.fetchall()

            produtos_originais_tuple = []
            produtos_modificados_tuple = []

            for i in dados_aux:
                data_t, cod_t, desc_t, ref_t, um_t, qtde_t, func_t = i

                qtde_float_t = valores_para_float(qtde_t)

                dados = (cod_t, qtde_float_t)
                produtos_originais_tuple.append(dados)

            for ii in dados_consumo:
                data_m, cod_m, desc_m, ref_m, um_m, qtde_m, func_m = ii

                qtde_float_m = valores_para_float(qtde_m)

                dados = (cod_m, qtde_float_m)
                produtos_modificados_tuple.append(dados)

            foi_excluidos = 0
            foi_inseridos = 0

            lista_saidas = set(produtos_originais_tuple).difference(set(produtos_modificados_tuple))
            if lista_saidas:
                for iii in lista_saidas:
                    cod_sai, qtde_sai = iii

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT id, etapas FROM produto where codigo = '{cod_sai}';")
                    selects_sai = cursor.fetchall()
                    id_prod_sai = selects_sai[0][0]

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT id, id_movimentacao FROM CONSUMO_TEMPORARIO_OP "
                                   f"where id_produto = {id_prod_sai} AND num_op = {numero_op};")
                    selects_sai = cursor.fetchall()
                    id_consumo_aux = selects_sai[0][0]
                    id_mov = selects_sai[0][1]

                    cursor = conecta.cursor()
                    cursor.execute(f"DELETE FROM CONSUMO_TEMPORARIO_OP "
                                   f"where id = {id_consumo_aux};")

                    cursor = conecta.cursor()
                    cursor.execute(f"DELETE FROM MOVIMENTACAO "
                                   f"where id = {id_mov};")

                    foi_excluidos += 1

            lista_entradas = set(produtos_modificados_tuple).difference(set(produtos_originais_tuple))
            if lista_entradas:
                for iiii in lista_entradas:
                    cod_ent, qtde_ent = iiii

                    for titi in dados_consumo:
                        cod_titi = titi[1]
                        func_titi = titi[6]

                        if cod_ent == cod_titi:
                            cursor = conecta.cursor()
                            cursor.execute(f"SELECT id, etapas FROM produto where codigo = '{cod_ent}';")
                            selects_ent = cursor.fetchall()
                            id_prod_ent = selects_ent[0][0]

                            cur = conecta.cursor()
                            cur.execute(
                                f"SELECT id, funcionario FROM funcionarios where funcionario = '{func_titi}';")
                            detalhes_func = cur.fetchall()
                            id_func, nome_func = detalhes_func[0]

                            cur = conecta.cursor()
                            cur.execute(f"SELECT id, descricao, nome_usuario FROM ENVIA_PC "
                                        f"where descricao = '{self.nome_computador}';")
                            select = cur.fetchall()
                            id_pc = select[0][0]

                            cursor = conecta.cursor()
                            cursor.execute("select GEN_ID(GEN_MOVIMENTACAO_ID,0) from rdb$database;")
                            ultimo_oc0 = cursor.fetchall()
                            ultimo_oc1 = ultimo_oc0[0]
                            ultimo_mov = int(ultimo_oc1[0]) + 1

                            cursor = conecta.cursor()
                            cursor.execute(f"Insert into MOVIMENTACAO (ID, PRODUTO, OBS, TIPO, QUANTIDADE, "
                                           f"DATA, CODIGO, FUNCIONARIO, LOCALESTOQUE) values "
                                           f"(GEN_ID(GEN_MOVIMENTACAO_ID,1), {id_prod_ent}, 'OP N {numero_op}', "
                                           f"'220', {qtde_ent}, '{data_consumo_certo}', {cod_ent}, {id_func}, 19);")

                            cursor = conecta.cursor()
                            cursor.execute(f"Insert into CONSUMO_TEMPORARIO_OP "
                                           f"(ID, DATA, NUM_OP, ID_PRODUTO, QTDE, ID_FUNCIONARIO, ID_MOVIMENTACAO, "
                                           f"ID_PC) "
                                           f"values (GEN_ID(GEN_CONSUMO_TEMPORARIO_OP_ID,1), '{data_consumo_certo}', "
                                           f"{numero_op}, {id_prod_ent}, {qtde_ent}, {id_func}, {ultimo_mov}, "
                                           f"{id_pc});")

                            foi_inseridos += 1

                            break

            if foi_inseridos or foi_excluidos:
                conecta.commit()

                self.mensagem_alerta(f"Consumo Auxiliar na OP Nº {numero_op} atualizado com sucesso!")

                self.reiniciando_tela()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gera_excel(self, caminho_pasta):
        try:
            from unidecode import unidecode
            from openpyxl import load_workbook
            import pandas as pd
            from openpyxl.styles import Side, Alignment, Border, Font, PatternFill
            from sympy import frac
            from pathlib import Path

            num_op = self.line_Num.text()

            data_hoje = date.today()
            data_certa = data_hoje.strftime("%d/%m/%Y")

            dados_tabela = extrair_tabela(self.table_Lista)

            if dados_tabela:
                d_um = []

                for tabi in dados_tabela:
                    data, cod, desc, ref, um, qtde, func = tabi

                    if "," in qtde:
                        qtdezinha_com_ponto = qtde.replace(',', '.')
                        qtdezinha_float = float(qtdezinha_com_ponto)
                    else:
                        qtdezinha_float = float(qtde)

                    dados = (data, cod, desc, ref, um, qtdezinha_float, func)
                    d_um.append(dados)

                df = pd.DataFrame(d_um, columns=['Data', 'Código', 'Descrição', 'Referência', 'UM', 'Qtde',
                                                 'Funcionário'])

                codigo_int = {'Código': int}
                df = df.astype(codigo_int)
                qtde_float = {'Qtde': float}
                df = df.astype(qtde_float)

                camino = os.path.join('..', 'arquivos', 'modelo excel', 'ci_consumo_temp_op.xlsx')
                caminho_arquivo = definir_caminho_arquivo(camino)

                book = load_workbook(caminho_arquivo)

                desktop = Path.home() / "Desktop"
                desk_str = str(desktop)
                nome_req = f'\Consumo OP {num_op}.xlsx'
                caminho = (desk_str + nome_req)

                writer = pd.ExcelWriter(caminho, engine='openpyxl')

                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                linhas_frame = df.shape[0]
                colunas_frame = df.shape[1]

                colunas_certas = colunas_frame + 1

                ws = book.active

                inicia = 10
                rows = range(inicia, inicia + linhas_frame)
                columns = range(1, colunas_certas)

                for row in rows:
                    for col in columns:
                        ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                                wrap_text=True)
                        ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                          right=Side(border_style='thin', color='00000000'),
                                                          top=Side(border_style='thin', color='00000000'),
                                                          bottom=Side(border_style='thin', color='00000000'),
                                                          diagonal=Side(border_style='thick', color='00000000'),
                                                          diagonal_direction=0,
                                                          outline=Side(border_style='thin', color='00000000'),
                                                          vertical=Side(border_style='thin', color='00000000'),
                                                          horizontal=Side(border_style='thin', color='00000000'))

                ws.merge_cells(f'B6:B6')
                top_left_cell = ws[f'B6']
                c = ws[f'B6']
                c.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=14, bold=False)
                top_left_cell.value = num_op

                ws.merge_cells(f'G6:G6')
                top_left_cell = ws[f'G6']
                c = ws[f'G6']
                c.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=14, bold=False)
                top_left_cell.value = data_certa

                df.to_excel(writer, 'Sheet1', startrow=9, startcol=0, header=False, index=False)

                writer.save()
                self.label_6.setText(f'Excel gerado com sucesso!')

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaConsumoTemporario()
    tela.show()
    qt.exec_()
