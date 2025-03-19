import sys
from banco_dados.conexao import conecta
from forms.tela_exp_alterar import *
from arquivos.chamar_arquivos import definir_caminho_arquivo
from banco_dados.controle_erros import grava_erro_banco
from comandos.lines import validador_decimal, validador_so_numeros
from comandos.tabelas import extrair_tabela, lanca_tabela, layout_cabec_tab
from comandos.telas import tamanho_aplicacao, icone, cor_fundo_tela
from comandos.conversores import valores_para_float, float_para_moeda_reais, moeda_reais_para_float
from comandos.excel import edita_alinhamento, edita_bordas, linhas_colunas_p_edicao, edita_fonte, \
    edita_preenchimento
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
from datetime import date, datetime
import inspect
import os
import socket

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.styles import Font, Border, Side, Alignment

import traceback

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4


class TelaExpAlterar(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        self.nome_computador = socket.gethostname()

        cor_fundo_tela(self)
        nome_arquivo_com_caminho = inspect.getframeinfo(inspect.currentframe()).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_vendas.png")
        tamanho_aplicacao(self)
        self.layout_tabela_ov(self.table_OV)
        self.layout_tabela_exp(self.table_Exp)

        self.processando = False

        self.table_OV.viewport().installEventFilter(self)

        self.line_Num_Exp.editingFinished.connect(self.verifica_line_num_exp)

        self.radio_Maquinas.toggled.connect(self.abre_widget_maquinas_logistica)
        self.radio_Logistica.toggled.connect(self.abre_widget_maquinas_logistica)

        self.btn_Adicionar_Item.clicked.connect(self.adicionar_um_item)
        self.btn_Adicionar_Todos.clicked.connect(self.lanca_todos_produtos_ov)

        self.btn_ExcluirItem.clicked.connect(self.excluir_item_tab)
        self.btn_ExcluirTudo.clicked.connect(self.acionamento_btn_excluir_tudo)

        self.btn_Salvar.clicked.connect(self.verifica_salvamento)

        self.btn_Limpar.clicked.connect(self.limpa_tudo)

        self.btn_Separar.clicked.connect (self.manipula_comeco)

        validador_so_numeros(self.line_Num_Exp)

        validador_decimal(self.line_Peso_Liquido, 9999999.000)
        validador_decimal(self.line_Peso_Bruto, 9999999.000)
        validador_so_numeros(self.line_Volume)

        self.definir_combo_funcionario()
        self.definir_combo_veiculo()
        self.definir_emissao()

        self.widget_Maquinas.setHidden(True)
        self.widget_Logistica.setHidden(True)

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            nome_funcao_trat = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def layout_tabela_ov(self, nome_tabela):
        try:
            layout_cabec_tab(nome_tabela)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def layout_tabela_exp(self, nome_tabela):
        try:
            layout_cabec_tab(nome_tabela)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_emissao(self):
        try:
            data_hoje = date.today()
            self.date_Emissao.setDate(data_hoje)
            self.date_Emissao.setReadOnly(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_combo_funcionario(self):
        try:
            tabela = []

            self.combo_Motorista.clear()
            tabela.append("")

            cur = conecta.cursor()
            cur.execute(f"SELECT id, funcionario FROM funcionarios where ativo = 'S' order by funcionario;")
            detalhes_func = cur.fetchall()

            for dadus in detalhes_func:
                ides, func = dadus
                tabela.append(func)

            self.combo_Motorista.addItems(tabela)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_combo_veiculo(self):
        try:
            tabela = []

            self.combo_Veiculo.clear()
            tabela.append("")

            cur = conecta.cursor()
            cur.execute(f"SELECT id, descricao FROM PLACA_VEICULO order by descricao;")
            detalhes_func = cur.fetchall()

            for dadus in detalhes_func:
                ides, veiculo = dadus
                tabela.append(veiculo)

            self.combo_Veiculo.addItems(tabela)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_line_num_exp(self):
        if not self.processando:
            try:
                self.processando = True

                self.table_OV.setRowCount(0)
                self.table_Exp.setRowCount(0)

                num_exp = self.line_Num_Exp.text()
                if num_exp:
                    if int(num_exp) == 0:
                        self.mensagem_alerta('O campo "Código" não pode ser "0"!')
                        self.line_Num_Exp.clear()
                    else:
                        self.verifica_sql_exp()

            except Exception as e:
                nome_funcao = inspect.currentframe().f_code.co_name
                exc_traceback = sys.exc_info()[2]
                self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

            finally:
                self.processando = False

    def verifica_sql_exp(self):
        try:
            num_exp = self.line_Num_Exp.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT * FROM ordemexpedicao where id = {num_exp};")
            detalhes = cursor.fetchall()
            if not detalhes:
                self.mensagem_alerta('Este número da Ordem Expedição não existe!')
                self.line_Num_Exp.clear()
            else:
                self.lanca_dados_exp()
                self.lanca_produtos_exp()
                self.dados_ov_aberto_com_cliente()
                self.soma_totais()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def dados_ov_aberto_com_cliente(self):
        try:
            cliente = self.line_Cliente.text()
            if cliente:
                clientetete = cliente.find(" - ")
                id_cliente = cliente[:clientetete]

                tabela_nova = []

                cursor = conecta.cursor()
                cursor.execute(f"SELECT oc.numero, prod.codigo, prod.descricao, "
                               f"prod.unidade, prodoc.quantidade, "
                               f"prodoc.produzido, prod.quantidade, prod.localizacao "
                               f"FROM PRODUTOORDEMCOMPRA as prodoc "
                               f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                               f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                               f"INNER JOIN clientes as cli ON oc.cliente = cli.id "
                               f"LEFT JOIN pedidointerno as ped ON prodoc.id_pedido = ped.id "
                               f"where prodoc.quantidade > prodoc.produzido "
                               f"and (prodoc.id_expedicao IS NULL or prodoc.id_expedicao = 0) "
                               f"and oc.status = 'A' "
                               f"and oc.entradasaida = 'S'"
                               f"and oc.cliente = {id_cliente};")
                dados_interno = cursor.fetchall()
                if dados_interno:
                    for i in dados_interno:
                        num_ov, cod, descr, um, qtde_total, qtde_entr, saldo, local = i

                        total_float = float(qtde_total)
                        entregue_float = float(qtde_entr)
                        saldo_float = float(saldo)

                        falta_ent = total_float - entregue_float
                        falta_arr = round(valores_para_float(falta_ent), 2)

                        desc_tot = f"{descr} ({total_float})"

                        if saldo_float >= falta_ent:
                            dados = (num_ov, cod, desc_tot, um, falta_arr, local, saldo_float)
                            tabela_nova.append(dados)
                if tabela_nova:
                    lista_de_listas_ordenada = sorted(tabela_nova, key=lambda x: x[1])
                    lanca_tabela(self.table_OV, lista_de_listas_ordenada)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_exp(self):
        try:
            num_exp = self.line_Num_Exp.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT exp.emissao, exp.id_cliente, cli.razao, exp.status, exp.peso_bruto, "
                           f"exp.peso_liquido, exp.volume, exp.id_veiculo, exp.id_funcionario, "
                           f"exp.altura, exp.largura, exp.comprimento "
                           f"from ordemexpedicao as exp "
                           f"INNER JOIN clientes as cli ON exp.id_cliente = cli.id "
                           f"where exp.id = {num_exp};")
            dados_interno = cursor.fetchall()

            (emissao, id_cliente, nome_cliente, status, bruto, liquido, volume, id_veiculo, id_func, altura, largura,
             compr) = dados_interno[0]

            self.date_Emissao.setDate(emissao)

            if status == "A":
                self.label_Status.setText("ABERTO")
                self.liberar_campos_exp()
            elif status == "B":
                self.label_Status.setText("BAIXADO")
                self.bloquear_campos_exp()
            else:
                self.label_Status.setText(status)

            texto_cliente = f"{id_cliente} - {nome_cliente}"

            self.line_Cliente.setText(texto_cliente)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_produtos_exp(self):
        try:
            lista_final = []

            num_exp = self.line_Num_Exp.text()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT oc.numero, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, "
                           f"prod.unidade, prodoc.quantidade, prodoc.unitario, COALESCE(prodoc.ipi, '') as ipi "
                           f"FROM PRODUTOORDEMCOMPRA as prodoc "
                           f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                           f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                           f"where prodoc.id_expedicao = {num_exp};")
            dados_produtos = cursor.fetchall()

            if dados_produtos:
                for i in dados_produtos:
                    num_ov_exp, cod_exp, descr_exp, ref_exp, um_exp, qtde, unit_exp, ipi_exp = i

                    if ipi_exp:
                        if float(ipi_exp) == 0:
                            ipi_str = ""
                        else:
                            ipi_str = ipi_exp
                    else:
                        ipi_str = ""

                    qtde_float = valores_para_float(qtde)
                    unit_float = valores_para_float(unit_exp)

                    if ipi_exp:
                        ipi_float = valores_para_float(ipi_exp)
                        if ipi_float > 0:
                            total_certo = qtde_float * ((unit_float * (ipi_float / 100)) + unit_float)
                        else:
                            total_certo = qtde_float * unit_float
                    else:
                        total_certo = qtde_float * unit_float

                    total_dois = ("%.2f" % total_certo)

                    unit_moeda = float_para_moeda_reais(unit_exp)
                    total_moeda = float_para_moeda_reais(total_dois)

                    dados = (num_ov_exp, cod_exp, descr_exp, um_exp, qtde_float, unit_moeda, ipi_str, total_moeda)
                    lista_final.append(dados)

            if lista_final:
                lanca_tabela(self.table_Exp, lista_final)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def bloquear_campos_exp(self):
        try:
            self.line_Cliente.setReadOnly(True)
            self.line_Peso_Bruto.setReadOnly(True)
            self.line_Peso_Liquido.setReadOnly(True)
            self.line_Volume.setReadOnly(True)

            self.radio_Maquinas.setEnabled(False)
            self.radio_Logistica.setEnabled(False)

            self.combo_Motorista.setEnabled(False)
            self.combo_Veiculo.setEnabled(False)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def liberar_campos_exp(self):
        try:
            self.line_Cliente.setReadOnly(False)
            self.line_Peso_Bruto.setReadOnly(False)
            self.line_Peso_Liquido.setReadOnly(False)
            self.line_Volume.setReadOnly(False)

            self.radio_Maquinas.setEnabled(True)
            self.radio_Logistica.setEnabled(True)

            self.combo_Motorista.setEnabled(True)
            self.combo_Veiculo.setEnabled(True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def abre_widget_maquinas_logistica(self):
        try:
            if self.radio_Maquinas.isChecked():
                self.widget_Maquinas.setHidden(False)
                self.widget_Logistica.setHidden(True)
                self.combo_Motorista.setFocus()
            else:
                self.widget_Maquinas.setHidden(True)
                self.widget_Logistica.setHidden(False)
                self.line_Altura.setFocus()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def eventFilter(self, source, event):
        try:
            qwidget_table = self.table_OV
            if (event.type() == QtCore.QEvent.MouseButtonDblClick and
                    event.buttons() == QtCore.Qt.LeftButton and
                    source is qwidget_table.viewport()):
                item = qwidget_table.currentItem()

                linha_selecao = self.table_OV.currentRow()

                extrai_recomendados = extrair_tabela(qwidget_table)
                item_selecionado = extrai_recomendados[item.row()]
                num_ov, cod, desc, um, falta, local, saldo = item_selecionado

                self.manipula_produtos_exp(num_ov, cod, falta, saldo, linha_selecao)

            return super(QMainWindow, self).eventFilter(source, event)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def adicionar_um_item(self):
        try:
            extrai_recomendados = extrair_tabela(self.table_OV)
            if not extrai_recomendados:
                self.mensagem_alerta(f'A tabela "Lista de Produtos" está vazia!')
            else:
                linha_selecao = self.table_OV.currentRow()

                dados_linha = []
                for coluna in range(self.table_OV.columnCount()):
                    item = self.table_OV.item(linha_selecao, coluna)
                    if item is not None:
                        dados_linha.append(item.text())

                if dados_linha:
                    num_ov, cod, desc, um, falta, local, saldo = dados_linha

                    self.manipula_produtos_exp(num_ov, cod, falta, saldo, linha_selecao)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_produtos_exp(self, num_ov, cod, qtde, saldo, linha_excluir):
        try:
            qtde_float = float(qtde)
            saldo_float = float(saldo)

            if qtde_float > saldo_float:
                qtde_ajustada = saldo
            else:
                qtde_ajustada = qtde

            cliente = self.line_Cliente.text()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            cursor = conecta.cursor()
            cursor.execute(f"SELECT oc.numero, prod.codigo, prod.descricao, "
                           f"COALESCE(prod.obs, '') as obs, "
                           f"prod.unidade, prodoc.unitario, COALESCE(prodoc.ipi, '') as ipi "
                           f"FROM PRODUTOORDEMCOMPRA as prodoc "
                           f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                           f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                           f"where oc.status = 'A' "
                           f"and oc.entradasaida = 'S'"
                           f"and oc.numero = {num_ov} "
                           f"and oc.cliente = {id_cliente} "
                           f"and prod.codigo = {cod};")
            dados_produtos = cursor.fetchall()

            extrai_produtos = extrair_tabela(self.table_Exp)

            num_pi, cod, descr, ref, um, unit, ipi = dados_produtos[0]

            if ipi:
                if float(ipi) == 0:
                    ipi_str = ""
                else:
                    ipi_str = ipi
            else:
                ipi_str = ""

            if qtde_ajustada and unit:
                qtde_float = valores_para_float(qtde_ajustada)
                unit_float = valores_para_float(unit)

                if ipi:
                    ipi_float = valores_para_float(ipi)

                    total_certo = qtde_float * ((unit_float * (ipi_float / 100)) + unit_float)
                else:
                    total_certo = qtde_float * unit_float
            else:
                total_certo = 0

            total_dois = ("%.2f" % total_certo)

            unit_moeda = float_para_moeda_reais(unit)
            total_moeda = float_para_moeda_reais(total_dois)

            dados = [num_ov, cod, descr, um, qtde_ajustada, unit_moeda, ipi_str, total_moeda]

            ja_existe = False
            for iii in extrai_produtos:
                num_ov_e, cod_e, des_e, um_e, qtde_e, unit_e, ipi_e, tot_e = iii

                if cod == cod_e and num_ov == num_ov_e:
                    ja_existe = True
                    break

            if not ja_existe:
                extrai_produtos.append(dados)

            if extrai_produtos:
                lanca_tabela(self.table_Exp, extrai_produtos)

            self.table_OV.removeRow(linha_excluir)
            self.soma_totais()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_todos_produtos_ov(self):
        try:
            cliente = self.line_Cliente.text()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            extrai_produtos_ov = extrair_tabela(self.table_OV)
            extrai_produtos = extrair_tabela(self.table_Exp)

            if extrai_produtos_ov:
                for prod_ov in extrai_produtos_ov:
                    num_ov, cod_ov, desc_ov, um_ov, falta_ov, local_ov, saldo_ov = prod_ov

                    qtde_float = float(falta_ov)
                    saldo_float = float(saldo_ov)

                    if qtde_float > saldo_float:
                        qtde_ajustada = saldo_ov
                    else:
                        qtde_ajustada = falta_ov

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT oc.numero, prod.codigo, prod.descricao, "
                                   f"COALESCE(prod.obs, '') as obs, "
                                   f"prod.unidade, prodoc.unitario, COALESCE(prodoc.ipi, '') as ipi "
                                   f"FROM PRODUTOORDEMCOMPRA as prodoc "
                                   f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                                   f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                                   f"where oc.status = 'A' "
                                   f"and oc.entradasaida = 'S'"
                                   f"and oc.numero = {num_ov} "
                                   f"and oc.cliente = {id_cliente} "
                                   f"and prod.codigo = {cod_ov};")
                    dados_produtos = cursor.fetchall()

                    num_ov_exp, cod_exp, descr_exp, ref_exp, um_exp, unit_exp, ipi_exp = dados_produtos[0]

                    if ipi_exp:
                        if float(ipi_exp) == 0:
                            ipi_str = ""
                        else:
                            ipi_str = ipi_exp
                    else:
                        ipi_str = ""

                    if qtde_ajustada and unit_exp:
                        qtde_float = valores_para_float(qtde_ajustada)
                        unit_float = valores_para_float(unit_exp)

                        if ipi_exp:
                            ipi_float = valores_para_float(ipi_exp)
                            if ipi_float > 0:
                                total_certo = qtde_float * ((unit_float * (ipi_float / 100)) + unit_float)
                            else:
                                total_certo = qtde_float * unit_float
                        else:
                            total_certo = qtde_float * unit_float
                    else:
                        total_certo = 0

                    total_dois = ("%.2f" % total_certo)

                    unit_moeda = float_para_moeda_reais(unit_exp)
                    total_moeda = float_para_moeda_reais(total_dois)

                    dados = [num_ov_exp, cod_exp, descr_exp, um_exp, qtde_ajustada, unit_moeda, ipi_str, total_moeda]

                    dados_str = [str(value).strip() for value in dados]
                    extrai_produtos_str = [[str(value).strip() for value in item[:2]] for item in extrai_produtos]

                    existe_dados = any(item == dados_str[:2] for item in extrai_produtos_str)

                    if not existe_dados:
                        extrai_produtos.append(dados)

            if extrai_produtos:
                lanca_tabela(self.table_Exp, extrai_produtos)

            self.table_OV.setRowCount(0)
            self.soma_totais()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item_tab(self):
        try:
            dados_linha = []

            nome_tabela = self.table_Exp

            dados_tab = extrair_tabela(nome_tabela)
            if not dados_tab:
                self.mensagem_alerta(f'A tabela "Produtos Expedição" está vazia!')
            else:
                linha = nome_tabela.currentRow()
                if linha >= 0:
                    dados_linha = []
                    for coluna in range(self.table_Exp.columnCount()):
                        item = self.table_Exp.item(linha, coluna)
                        if item is not None:
                            dados_linha.append(item.text())

                    nome_tabela.removeRow(linha)

            if dados_linha:
                self.devolve_item_tabela_exp(dados_linha)

            self.soma_totais()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def devolve_item_tabela_exp(self, dados_linha):
        try:
            num_ov, cod, desc, um, qtde, unit, ipi, total = dados_linha

            extrai_estrutura = extrair_tabela(self.table_OV)

            ja_existe = False
            for itens in extrai_estrutura:
                num_ov_con = itens[0]
                cod_con = itens[1]
                if cod_con == cod and num_ov_con == num_ov:
                    ja_existe = True
                    break

            if not ja_existe:
                cliente = self.line_Cliente.text()
                if cliente:
                    clientetete = cliente.find(" - ")
                    id_cliente = cliente[:clientetete]

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT oc.numero, prod.codigo, prod.descricao, "
                                   f"prod.unidade, prodoc.quantidade, "
                                   f"prodoc.produzido, prod.quantidade, prod.localizacao "
                                   f"FROM PRODUTOORDEMCOMPRA as prodoc "
                                   f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                                   f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                                   f"INNER JOIN clientes as cli ON oc.cliente = cli.id "
                                   f"LEFT JOIN pedidointerno as ped ON prodoc.id_pedido = ped.id "
                                   f"where oc.status = 'A' "
                                   f"and oc.entradasaida = 'S'"
                                   f"and oc.cliente = {id_cliente} "
                                   f"and oc.numero = {num_ov} "
                                   f"and prod.codigo = {cod};")
                    dados_interno = cursor.fetchall()

                    if dados_interno:
                        for i in dados_interno:
                            num_ov, cod, descr, um, qtde_total, qtde_entr, saldo, local = i

                            total_float = float(qtde_total)
                            entregue_float = float(qtde_entr)
                            saldo_float = float(saldo)

                            falta_ent = total_float - entregue_float
                            falta_arr = round(valores_para_float(falta_ent), 2)

                            desc_tot = f"{descr} ({total_float})"

                            if saldo_float >= falta_ent:
                                dados = (num_ov, cod, desc_tot, um, falta_arr, local, saldo_float)
                                extrai_estrutura.append(dados)
                    if extrai_estrutura:
                        lista_de_listas_ordenada = sorted(extrai_estrutura, key=lambda x: x[1])
                        lanca_tabela(self.table_OV, lista_de_listas_ordenada)

            else:
                self.mensagem_alerta("Este produto já foi adicionado!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def acionamento_btn_excluir_tudo(self):
        try:
            tabela_nova = extrair_tabela(self.table_OV)

            extrai_pedido = extrair_tabela(self.table_Exp)

            if extrai_pedido:
                for itens in extrai_pedido:
                    num_ov, cod, descr, um, qtde, unit, ipi, total = itens

                    extrai_estrutura = extrair_tabela(self.table_OV)

                    ja_existe = False
                    for itenss in extrai_estrutura:
                        num_ov_con = itenss[0]
                        cod_con = itenss[1]
                        if cod_con == cod and num_ov_con == num_ov:
                            ja_existe = True
                            break

                    if not ja_existe:
                        cliente = self.line_Cliente.text()
                        if cliente:
                            clientetete = cliente.find(" - ")
                            id_cliente = cliente[:clientetete]

                            cursor = conecta.cursor()
                            cursor.execute(f"SELECT oc.numero, prod.codigo, prod.descricao, "
                                           f"prod.unidade, prodoc.quantidade, "
                                           f"prodoc.produzido, prod.quantidade, prod.localizacao "
                                           f"FROM PRODUTOORDEMCOMPRA as prodoc "
                                           f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                                           f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                                           f"INNER JOIN clientes as cli ON oc.cliente = cli.id "
                                           f"LEFT JOIN pedidointerno as ped ON prodoc.id_pedido = ped.id "
                                           f"where oc.status = 'A' "
                                           f"and oc.entradasaida = 'S'"
                                           f"and oc.cliente = {id_cliente} "
                                           f"and oc.numero = {num_ov} "
                                           f"and prod.codigo = {cod};")
                            dados_interno = cursor.fetchall()
                            if dados_interno:
                                for i in dados_interno:
                                    num_ov, cod, descr, um, qtde_total, qtde_entr, saldo, local = i

                                    total_float = float(qtde_total)
                                    entregue_float = float(qtde_entr)
                                    saldo_float = float(saldo)

                                    falta_ent = total_float - entregue_float
                                    falta_arr = round(valores_para_float(falta_ent), 2)

                                    desc_tot = f"{descr} ({total_float})"

                                    if saldo_float >= falta_ent:
                                        dados = (num_ov, cod, desc_tot, um, falta_arr, local, saldo_float)
                                        tabela_nova.append(dados)

            if tabela_nova:
                lista_de_listas_ordenada = sorted(tabela_nova, key=lambda x: x[1])
                lanca_tabela(self.table_OV, lista_de_listas_ordenada)

            self.table_Exp.setRowCount(0)
            self.soma_totais()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def soma_totais(self):
        try:
            soma_total_geral = 0.00

            extrai_produtos = extrair_tabela(self.table_Exp)
            if extrai_produtos:
                for i in extrai_produtos:
                    total = i[7]

                    total_float = moeda_reais_para_float(total)
                    soma_total_geral += total_float

            total_geral = str("%.2f" % soma_total_geral)
            total_geral_moeda = float_para_moeda_reais(total_geral)
            self.line_Total_Geral.setText(total_geral_moeda)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tudo(self):
        self.line_Cliente.clear()

        self.label_separa.setText("")

        self.table_OV.setRowCount(0)
        self.table_Exp.setRowCount(0)

        self.line_Total_Geral.clear()

        self.line_Peso_Bruto.clear()
        self.line_Peso_Liquido.clear()
        self.line_Volume.clear()

        self.line_Largura.clear()
        self.line_Altura.clear()
        self.line_Comprimento.clear()

        self.radio_Maquinas.setChecked(False)
        self.radio_Logistica.setChecked(False)
        self.widget_Maquinas.setHidden(True)
        self.widget_Logistica.setHidden(True)

        self.definir_emissao()
        self.definir_combo_funcionario()
        self.definir_combo_veiculo()

    def manipula_comeco(self):
        try:
            lista_email = []

            cliente = self.line_Cliente.text()

            extrai_pedido = extrair_tabela(self.table_Exp)

            if extrai_pedido and cliente:
                clientetete = cliente.find(" - ")
                id_cliente = cliente[:clientetete]

                for itens in extrai_pedido:
                    num_ov, codigo, descr, um, qtde, unit, ipi, total = itens

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prodoc.id, prod.codigo, prod.descricao, "
                                   f"COALESCE(prod.obs, ''), COALESCE(prod.localizacao, ''), prod.quantidade, "
                                   f"prod.unidade, prodoc.unitario, COALESCE(prodoc.ipi, ''), prodoc.produzido "
                                   f"FROM PRODUTOORDEMCOMPRA as prodoc "
                                   f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                                   f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                                   f"where oc.status = 'A' "
                                   f"and oc.entradasaida = 'S'"
                                   f"and oc.numero = {num_ov} "
                                   f"and oc.cliente = {id_cliente} "
                                   f"and prod.codigo = {codigo};")
                    dados_produtos = cursor.fetchall()
                    ref = dados_produtos[0][3]
                    local = dados_produtos[0][4]
                    saldo = dados_produtos[0][5]
                    produzido = dados_produtos[0][9]

                    falta = "%.3f" % (float(qtde) - float(produzido))

                    dadus = (codigo, descr, ref, um, falta, local, saldo)
                    lista_email.append(dadus)

            if lista_email:

                caminho = fr'C:\Users\Anderson\Desktop\Listagem - OV.pdf'

                self.gerar_pdf_listagem_separar(caminho, lista_email)

                self.label_separa.setText("PDF criado com sucesso!")

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def adicionar_tabelas_listagem(self, dados, cabecalho):
        try:
            elements = []

            style_lista = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.gray),
                                      ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                                      ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                      ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                                      ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                                      ('GRID', (0, 0), (-1, -1), 1, colors.black),
                                      ('FONTSIZE', (0, 0), (-1, 0), 10),
                                      ('FONTSIZE', (0, 1), (-1, -1), 8)])

            table = Table([cabecalho] + dados)
            table.setStyle(style_lista)
            elements.append(table)

            return elements

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gerar_pdf_listagem_separar(self, caminho_listagem, produtos_ov):
        try:
            margem_esquerda = 0
            margem_direita = 5
            margem_superior = 25
            margem_inferior = 5

            data_hoje = date.today()

            cliente = self.line_Cliente.text()

            dados_cliente = [(data_hoje, cliente)]

            doc = SimpleDocTemplate(caminho_listagem, pagesize=A4,
                                    leftMargin=margem_esquerda,
                                    rightMargin=margem_direita,
                                    topMargin=margem_superior,
                                    bottomMargin=margem_inferior)

            cabecalho_op = ['DATA', 'CLIENTE']
            elem_op = self.adicionar_tabelas_listagem(dados_cliente, cabecalho_op)

            cabecalho_lista = ['CÓDIGO', 'DESCRIÇÃO', 'REFERÊNCIA', 'UM', 'QTDE', 'LOCALIZAÇÃO', 'SALDO']
            elem_lista = self.adicionar_tabelas_listagem(produtos_ov, cabecalho_lista)

            cabecalho_transp = ['', 'TRANSPORTE']
            dados_transp = [('PESO LÍQUIDO', ''), ('PESO BRUTO', ''), ('VOLUME', '')]
            elem_transp = self.adicionar_tabelas_listagem(dados_transp, cabecalho_transp)

            cabecalho_medida = ['MEDIDAS', '            ']
            dados_medida = [('ALTURA (MM)', ''), ('LARGURA (MM)', ''), ('COMPRIMENTO (MM)', '')]
            elem_medida = self.adicionar_tabelas_listagem(dados_medida, cabecalho_medida)

            cabecalho_motorista = ['DAMDFE', 'MOTORISTA']
            dados_motorista = [('PLACA', ''), ('NOME', ''), ('CPF', '')]
            elem_motorista = self.adicionar_tabelas_listagem(dados_motorista, cabecalho_motorista)

            espaco_em_branco = Table([[None]], style=[('SIZE', (0, 0), (0, 0), 20)])

            # Criar tabela para colocar medidas e motorista lado a lado
            tabela_medida_motorista = Table([[elem_transp, elem_medida, elem_motorista]],
                                            colWidths=[170, 170])  # Ajuste as larguras conforme necessário

            elementos = (elem_op + [espaco_em_branco] +
                         elem_lista + [espaco_em_branco] +
                         [tabela_medida_motorista])  # Adiciona a tabela com medidas e motorista lado a lado

            doc.build(elementos)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_salvamento(self):
        try:
            extrai_pedido = extrair_tabela(self.table_Exp)
            num_exp = self.line_Num_Exp.text()
            cliente = self.line_Cliente.text()
            peso_bruto = self.line_Peso_Bruto.text()
            peso_liquido = self.line_Peso_Liquido.text()
            volume = self.line_Volume.text()

            motorista = self.combo_Motorista.currentText()
            veiculo = self.combo_Veiculo.currentText()

            if not extrai_pedido:
                self.mensagem_alerta(f'A tabela "Produtos Expedição" está vazia!')
            elif not num_exp:
                self.mensagem_alerta(f'O campo "Nº EXP" não pode estar vazio!')
            elif not cliente:
                self.mensagem_alerta(f'O campo "Cliente" não pode estar vazio!')
            elif not peso_bruto:
                self.mensagem_alerta(f'O campo "Peso Bruto" não pode estar vazio!')
            elif not peso_liquido:
                self.mensagem_alerta(f'O campo "Peso Líq." não pode estar vazio!')
            elif not volume:
                self.mensagem_alerta(f'O campo "Volume" não pode estar vazio!')
            elif not self.radio_Maquinas.isChecked() and not self.radio_Logistica.isChecked():
                self.mensagem_alerta(f'Defina o responsável pela definição do transporte!.')
            else:
                if self.radio_Maquinas.isChecked():
                    if not veiculo:
                        self.mensagem_alerta(f'O campo "Veículo" não pode estar vazio!')
                    elif not motorista:
                        self.mensagem_alerta(f'O campo "Motorista" não pode estar vazio!')
                    else:
                        self.salvar_expedicao_maq()

                elif self.radio_Logistica.isChecked():
                    if not self.line_Altura.text():
                        self.mensagem_alerta(f'O campo "Altura" não pode estar vazio!')
                    elif not self.line_Largura.text():
                        self.mensagem_alerta(f'O campo "Largura" não pode estar vazio!')
                    elif not self.line_Comprimento.text():
                        self.mensagem_alerta(f'O campo "Comprimento" não pode estar vazio!')
                    else:
                        self.salvar_expedicao_log()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_expedicao_log(self):
        try:
            print("salvar")
            cliente = self.line_Cliente.text()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            num_exp = self.line_Num_Exp.text()

            peso_bruto = self.line_Peso_Bruto.text()
            peso_bruto_float = valores_para_float(peso_bruto)

            peso_liquido = self.line_Peso_Liquido.text()
            peso_liquido_float = valores_para_float(peso_liquido)

            volume = self.line_Volume.text()

            altura = self.line_Altura.text()
            largura = self.line_Largura.text()
            compr = self.line_Comprimento.text()

            datamov = self.date_Emissao.text()
            date_mov = datetime.strptime(datamov, '%d/%m/%Y').date()
            data_mov_certa = str(date_mov)

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ORDEMEXPEDICAO (ID, EMISSAO, ID_CLIENTE, PESO_BRUTO, PESO_LIQUIDO, "
                           f"VOLUME, ALTURA, LARGURA, COMPRIMENTO, NOME_PC, STATUS) "
                           f"values (GEN_ID(GEN_ORDEMEXPEDICAO_ID,1), '{data_mov_certa}', "
                           f"'{id_cliente}', '{peso_bruto_float}', '{peso_liquido_float}', {volume}, '{altura}', "
                           f"'{largura}', '{compr}',  '{self.nome_computador}', 'A');")

            extrai_pedido = extrair_tabela(self.table_Exp)

            for itens in extrai_pedido:
                num_ov = itens[0]
                cod = itens[1]

                cursor = conecta.cursor()
                cursor.execute(f"SELECT prodoc.id, prod.codigo, prod.descricao, "
                               f"COALESCE(prod.obs, '') as obs, "
                               f"prod.unidade, prodoc.unitario, COALESCE(prodoc.ipi, '') as ipi "
                               f"FROM PRODUTOORDEMCOMPRA as prodoc "
                               f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                               f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                               f"where oc.status = 'A' "
                               f"and oc.entradasaida = 'S'"
                               f"and oc.numero = {num_ov} "
                               f"and oc.cliente = {id_cliente} "
                               f"and prod.codigo = {cod};")
                dados_produtos = cursor.fetchall()
                id_prodoc = dados_produtos[0][0]

                cursor = conecta.cursor()
                cursor.execute(f"UPDATE PRODUTOORDEMCOMPRA SET id_expedicao = {num_exp} "
                               f"WHERE id = {id_prodoc};")

            conecta.commit()

            print("salvado")

            self.gera_excel()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_expedicao_maq(self):
        try:
            print("salvar")
            cliente = self.line_Cliente.text()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            num_exp = self.line_Num_Exp.text()

            peso_bruto = self.line_Peso_Bruto.text()
            peso_bruto_float = valores_para_float(peso_bruto)

            peso_liquido = self.line_Peso_Liquido.text()
            peso_liquido_float = valores_para_float(peso_liquido)

            volume = self.line_Volume.text()

            motorista = self.combo_Motorista.currentText()
            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, funcionario FROM FUNCIONARIOS where funcionario = '{motorista}';")
            dados_funcionario = cursor.fetchall()
            id_funcionario = dados_funcionario[0][0]

            veiculo = self.combo_Veiculo.currentText()
            cursor = conecta.cursor()
            cursor.execute(f"SELECT id, placa FROM PLACA_VEICULO where descricao = '{veiculo}';")
            dados_veiculo = cursor.fetchall()
            id_veiculo = dados_veiculo[0][0]

            datamov = self.date_Emissao.text()
            date_mov = datetime.strptime(datamov, '%d/%m/%Y').date()
            data_mov_certa = str(date_mov)

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ORDEMEXPEDICAO (ID, EMISSAO, ID_CLIENTE, PESO_BRUTO, PESO_LIQUIDO, "
                           f"VOLUME, ID_VEICULO, ID_FUNCIONARIO, NOME_PC) "
                           f"values (GEN_ID(GEN_ORDEMEXPEDICAO_ID,1), '{data_mov_certa}', "
                           f"'{id_cliente}', '{peso_bruto_float}', '{peso_liquido_float}', {volume}, {id_veiculo}, "
                           f"{id_funcionario}, '{self.nome_computador}');")

            extrai_pedido = extrair_tabela(self.table_Exp)

            for itens in extrai_pedido:
                num_ov = itens[0]
                cod = itens[1]

                cursor = conecta.cursor()
                cursor.execute(f"SELECT prodoc.id, prod.codigo, prod.descricao, "
                               f"COALESCE(prod.obs, '') as obs, "
                               f"prod.unidade, prodoc.unitario, COALESCE(prodoc.ipi, '') as ipi "
                               f"FROM PRODUTOORDEMCOMPRA as prodoc "
                               f"INNER JOIN produto as prod ON prodoc.produto = prod.id "
                               f"INNER JOIN ordemcompra as oc ON prodoc.mestre = oc.id "
                               f"where oc.status = 'A' "
                               f"and oc.entradasaida = 'S'"
                               f"and oc.numero = {num_ov} "
                               f"and oc.cliente = {id_cliente} "
                               f"and prod.codigo = {cod};")
                dados_produtos = cursor.fetchall()
                id_prodoc = dados_produtos[0][0]

                cursor = conecta.cursor()
                cursor.execute(f"UPDATE PRODUTOORDEMCOMPRA SET id_expedicao = {num_exp} "
                               f"WHERE id = {id_prodoc};")

            conecta.commit()

            print("salvado")

            self.gera_excel()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gera_excel(self):
        try:
            tem_s = 0
            tem_n = 0

            cliente = self.line_Cliente.text()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            cursor = conecta.cursor()
            cursor.execute(f"SELECT registro, razao, estado FROM clientes where id = {id_cliente};")
            dados_cliente = cursor.fetchall()

            registro, razao, estado_cli = dados_cliente[0]
            texto_cliente = f"{registro} - {razao}"

            num_exp = self.line_Num_Exp.text()

            peso_bruto = self.line_Peso_Bruto.text()
            txt_peso_bruto = f"{peso_bruto} KG"

            peso_liquido = self.line_Peso_Liquido.text()
            txt_peso_liquido = f"{peso_liquido} KG"

            volume = self.line_Volume.text()

            if self.radio_Maquinas.isChecked():
                veiculo = self.combo_Veiculo.currentText()
                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, placa FROM PLACA_VEICULO where descricao = '{veiculo}';")
                dados_veiculo = cursor.fetchall()

                placa = dados_veiculo[0][1]

                txt_transp = f"PLACA: {placa}"
            else:
                altura = self.line_Altura.text()
                largura = self.line_Largura.text()
                compr = self.line_Comprimento.text()

                txt_transp = f"CFE. LOGÍSTICA " \
                             f"(ALTURA: {altura} MM / " \
                             f"LARGURA: {largura} MM / " \
                             f"COMPRIMENTO: {compr} MM)"

            dados_tabela1 = []
            dados_tabela = extrair_tabela(self.table_Exp)

            for index, iii in enumerate(dados_tabela):
                seq = index + 1
                num_ov, cod, desc, um, qtde, unit, ipi, total = iii
                didi = (seq, num_ov, cod, desc, um, qtde, unit, ipi, total)
                dados_tabela1.append(didi)

            dados_p_descricao = []

            camino = os.path.join('..', 'arquivos', 'modelo excel', 'exp_incluir.xlsx')
            caminho_arquivo = definir_caminho_arquivo(camino)

            book = load_workbook(caminho_arquivo)

            desktop = Path.home() / "Desktop"
            desk_str = str(desktop)
            nome_req = f'\Solicitação NF {razao} Nº {num_exp}.xlsx'
            caminho = (desk_str + nome_req)

            writer = pd.ExcelWriter(caminho, engine='openpyxl')

            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            ws = book.active

            dados_um = []
            dados_dois = []
            dados_tres = []

            texto_ocs = ""
            numeros_adicionados = set()

            for tabi in dados_tabela1:
                dados_p_descricao.append(tabi)
                seq, num_ov, cod, desc, um, qtde, unit, ipi, total = tabi

                if num_ov not in numeros_adicionados:
                    if texto_ocs:
                        texto_ocs += " / "
                    texto_ocs += f"OC {num_ov}"
                    numeros_adicionados.add(num_ov)

                if not ipi:
                    indust = "NÃO"
                elif ipi == 0.00:
                    indust = "NÃO"
                else:
                    indust = "SIM"

                if "," in qtde:
                    qtdezinha_com_ponto = qtde.replace(',', '.')
                    qtdezinha_float = float(qtdezinha_com_ponto)
                else:
                    qtdezinha_float = float(qtde)

                if indust == "SIM":
                    tem_s += 1
                if indust == "NÃO":
                    tem_n += 1

                dados1 = (cod, qtdezinha_float)
                dados2 = (um, indust, unit)

                dados_um.append(dados1)
                dados_dois.append(dados2)
                dados_tres.append(total)

            inicio_produtos = 21

            linhas_produtos = len(dados_tabela1)
            if linhas_produtos < 10:
                dif = 10 - linhas_produtos
                totalzao = dif + linhas_produtos
                dedos = ['', '', '', '', '', '', '', '', '']
                for repite in range(dif):
                    dados_p_descricao.append(dedos)
            else:
                totalzao = linhas_produtos

            numero_ov = [tabi[1] for tabi in dados_tabela1]
            df_ordem = pd.DataFrame({'OC': numero_ov})
            oc_int = {'OC': int}
            df_ordem = df_ordem.astype(oc_int)
            df_ordem.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=0, header=False, index=False)

            num_seq = [tabi[0] for tabi in dados_tabela1]
            df_ordem = pd.DataFrame({'ITEM': num_seq})
            seq_int = {'ITEM': int}
            df_sequencia = df_ordem.astype(seq_int)
            df_sequencia.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=1, header=False,
                                  index=False)

            self.manipula_do_cod_e_qtde(writer, inicio_produtos, dados_um)
            self.manipula_da_um_ate_unid(writer, inicio_produtos, dados_dois)
            self.manipula_total(writer, inicio_produtos, dados_tres)
            self.manipula_descricao(ws, writer, inicio_produtos, dados_p_descricao)

            for cell in linhas_colunas_p_edicao(ws, inicio_produtos, inicio_produtos + totalzao - 1, 2, 14):
                edita_bordas(cell)
                edita_alinhamento(cell, ali_vertical='bottom')
                edita_fonte(cell)

            for cell in linhas_colunas_p_edicao(ws, inicio_produtos, inicio_produtos + totalzao - 1, 1, 1):
                edita_alinhamento(cell, ali_vertical='bottom')
                edita_fonte(cell)

                cell.border = Border(right=Side(border_style='medium', color='00000000'))

            for cell in linhas_colunas_p_edicao(ws, inicio_produtos, inicio_produtos + totalzao - 1, 14, 14):
                edita_preenchimento(cell)
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='medium', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'),
                                     diagonal=Side(border_style='thick', color='00000000'),
                                     diagonal_direction=0,
                                     outline=Side(border_style='thin', color='00000000'),
                                     vertical=Side(border_style='thin', color='00000000'),
                                     horizontal=Side(border_style='thin', color='00000000'))

            altura_celula = 24.75
            for linha in range(inicio_produtos, inicio_produtos + totalzao):
                ws.row_dimensions[linha].height = altura_celula

            personalizacao = ['Times New Roman', 10, False, 'center', 'bottom']
            self.lanca_dados_mesclado(ws, 'D14:N14', 'D14', txt_transp, personalizacao)

            personalizacao = ['Times New Roman', 10, False, 'left', 'bottom']
            self.lanca_dados_mesclado(ws, 'C11:N11', 'C11', texto_cliente, personalizacao)

            texto_operacao = ""

            if tem_s and tem_n:
                if estado_cli == "RS":
                    texto_operacao += "5101 - VENDA PROD. ESTAB. / 5102 - VENDA MERCADORIAS"
                else:
                    texto_operacao += "6101 - VENDA PROD. ESTAB. (F.E.) / 6102 - VENDA MERCADORIAS (F.E.)"
            elif tem_s:
                if estado_cli == "RS":
                    texto_operacao += "5101 - VENDA PROD. ESTAB."
                else:
                    texto_operacao += "6101 - VENDA PROD. ESTAB. (F.E.)"
            else:
                if estado_cli == "RS":
                    texto_operacao += "5102 - VENDA MERCADORIAS"
                else:
                    texto_operacao += "6102 - VENDA MERCADORIAS (F.E.)"

            if tem_s:
                personalizacao = ['Times New Roman', 10, True, 'center', 'center']
                informacao = "X"
                self.lanca_dados_coluna(ws, "E5", informacao, personalizacao)

            if tem_n:
                personalizacao = ['Times New Roman', 10, True, 'center', 'center']
                informacao = "X"
                self.lanca_dados_coluna(ws, "E7", informacao, personalizacao)

            personalizacao = ['Times New Roman', 10, True, 'left', 'bottom']
            self.lanca_dados_mesclado(ws, 'E16:N16', 'E16', texto_operacao, personalizacao)

            personalizacao = ['Times New Roman', 10, False, 'left', 'bottom']
            self.lanca_dados_mesclado(ws, 'C13:N13', 'C13', texto_ocs, personalizacao)

            linha_vazia = inicio_produtos + totalzao
            for linha in range(linha_vazia, linha_vazia + 1):
                ws.row_dimensions[linha].height = altura_celula

                ws.merge_cells(f'C{linha}:M{linha}')
                celula_sup_esq = ws[f'C{linha}']
                celula_sup_esq.value = ""
                for cell in linhas_colunas_p_edicao(ws, linha, linha, 3, 13):
                    cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                         right=Side(border_style='thin', color='00000000'),
                                         top=Side(border_style='medium', color='00000000'),
                                         bottom=Side(border_style='thin', color='00000000'),
                                         diagonal=Side(border_style='thick', color='00000000'),
                                         diagonal_direction=0,
                                         outline=Side(border_style='thin', color='00000000'),
                                         vertical=Side(border_style='thin', color='00000000'),
                                         horizontal=Side(border_style='thin', color='00000000'))

                for cell in linhas_colunas_p_edicao(ws, linha, linha, 2, 2):
                    cell.border = Border(top=Side(border_style='medium', color='00000000'))

                for cell in linhas_colunas_p_edicao(ws, linha, linha, 14, 14):
                    cell.border = Border(top=Side(border_style='medium', color='00000000'))

            segunda_linha_vazia = inicio_produtos + totalzao + 1
            for linha in range(segunda_linha_vazia, segunda_linha_vazia + 1):
                ws.row_dimensions[linha].height = altura_celula

                ws.merge_cells(f'C{linha}:M{linha}')
                celula_sup_esq = ws[f'C{linha}']
                celula_sup_esq.value = ""
                for cell in linhas_colunas_p_edicao(ws, linha, linha, 3, 13):
                    edita_bordas(cell)

            segunda_linha_vazia = inicio_produtos + totalzao + 2
            for linha in range(segunda_linha_vazia, segunda_linha_vazia + 1):
                ws.row_dimensions[linha].height = altura_celula

                ws.merge_cells(f'C{linha}:H{linha}')
                celula_sup_esq = ws[f'C{linha}']
                celula_sup_esq.value = "* Produtos com * cod informados se possível."
                for cell in linhas_colunas_p_edicao(ws, linha, linha, 3, 8):
                    edita_alinhamento(cell, ali_vertical='bottom')
                    edita_fonte(cell)
                    edita_bordas(cell)

            segunda_linha_vazia = inicio_produtos + totalzao + 3
            for linha in range(segunda_linha_vazia, segunda_linha_vazia + 1):
                ws.row_dimensions[linha].height = altura_celula

                ws.merge_cells(f'C{linha}:H{linha}')
                celula_sup_esq = ws[f'C{linha}']
                celula_sup_esq.value = "* Cada nota possui somente 18 linhas disponíveis."
                for cell in linhas_colunas_p_edicao(ws, linha, linha, 3, 8):
                    edita_alinhamento(cell, ali_vertical='bottom')
                    edita_fonte(cell)
                    edita_bordas(cell)

            segunda_linha_vazia = inicio_produtos + totalzao + 4
            for linha in range(segunda_linha_vazia, segunda_linha_vazia + 1):
                ws.row_dimensions[linha].height = altura_celula

                ws.merge_cells(f'C{linha}:H{linha}')
                celula_sup_esq = ws[f'C{linha}']
                celula_sup_esq.value = "* Favor digitar volume e peso no final de cada nota."
                for cell in linhas_colunas_p_edicao(ws, linha, linha, 3, 8):
                    edita_alinhamento(cell, ali_vertical='bottom')
                    edita_fonte(cell)
                    edita_bordas(cell)

            linha_bruto = inicio_produtos + totalzao + 2
            person = ['Times New Roman', 10, False, 'center', 'bottom']
            self.lanca_dados_mesclado(ws, f'K{linha_bruto}:M{linha_bruto}', f'K{linha_bruto}', txt_peso_bruto, person)
            for cell in linhas_colunas_p_edicao(ws, linha_bruto, linha_bruto, 11, 13):
                edita_bordas(cell)

            personalizacao = ['Times New Roman', 10, True, 'center', 'bottom']
            informacao = "PESO BRUTO"
            self.lanca_dados_coluna(ws, f'J{linha_bruto}', informacao, personalizacao)
            for cell in linhas_colunas_p_edicao(ws, linha_bruto, linha_bruto, 10, 10):
                edita_bordas(cell)

            linha_liq = inicio_produtos + totalzao + 3
            person = ['Times New Roman', 10, False, 'center', 'bottom']
            self.lanca_dados_mesclado(ws, f'K{linha_liq}:M{linha_liq}', f'K{linha_liq}', txt_peso_liquido, person)
            for cell in linhas_colunas_p_edicao(ws, linha_liq, linha_liq, 11, 13):
                edita_bordas(cell)

            personalizacao = ['Times New Roman', 10, True, 'center', 'bottom']
            informacao = "PESO LÍQUIDO"
            self.lanca_dados_coluna(ws, f'J{linha_liq}', informacao, personalizacao)
            for cell in linhas_colunas_p_edicao(ws, linha_liq, linha_liq, 10, 10):
                edita_bordas(cell)

            linha_vol = inicio_produtos + totalzao + 4
            person = ['Times New Roman', 10, False, 'center', 'bottom']
            self.lanca_dados_mesclado(ws, f'K{linha_vol}:M{linha_vol}', f'K{linha_vol}', volume, person)
            for cell in linhas_colunas_p_edicao(ws, linha_vol, linha_vol, 11, 13):
                edita_bordas(cell)

            personalizacao = ['Times New Roman', 10, True, 'center', 'bottom']
            informacao = "VOLUME"
            self.lanca_dados_coluna(ws, f'J{linha_vol}', informacao, personalizacao)
            for cell in linhas_colunas_p_edicao(ws, linha_vol, linha_vol, 10, 10):
                edita_bordas(cell)

            writer.save()

            self.mensagem_alerta(f'Num {num_exp} criada e excel gerado com sucesso!')
            self.limpa_tudo()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_mesclado(self, bloco_book, mesclado, celula, informacao, fonte_alinhamento):
        try:
            nome_fonte, tamanho_fonte, e_negrito, alin_hor, alin_ver = fonte_alinhamento

            bloco_book.merge_cells(mesclado)
            celula_sup_esq = bloco_book[celula]
            cel = bloco_book[celula]
            cel.alignment = Alignment(horizontal=alin_hor, vertical=alin_ver, text_rotation=0,
                                      wrap_text=False, shrink_to_fit=False, indent=0)
            cel.font = Font(name=nome_fonte, size=tamanho_fonte, bold=e_negrito)
            celula_sup_esq.value = informacao

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_dados_coluna(self, bloco_book, celula, informacao, fonte_alinhamento):
        try:
            nome_fonte, tamanho_fonte, e_negrito, alin_hor, alin_ver = fonte_alinhamento

            celula_sup_esq = bloco_book[celula]
            cel = bloco_book[celula]
            cel.alignment = Alignment(horizontal=alin_hor, vertical=alin_ver, text_rotation=0,
                                      wrap_text=False, shrink_to_fit=False, indent=0)
            cel.font = Font(name=nome_fonte, size=tamanho_fonte, bold=e_negrito)
            celula_sup_esq.value = informacao

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def mesclar_descricao(self, ws, desc_produtos, ini_prod):
        try:
            startcol_descricao = 5
            endcol_descricao = 10

            for idx, desc in enumerate(desc_produtos):
                row_num = ini_prod + idx

                ws.cell(row=row_num, column=startcol_descricao).value = desc
                ws.cell(row=row_num, column=startcol_descricao).alignment = Alignment(horizontal='center',
                                                                                      vertical='bottom', wrap_text=True)

                if len(desc_produtos) > 1:
                    ws.merge_cells(start_row=row_num, start_column=startcol_descricao, end_row=row_num,
                                   end_column=endcol_descricao)

            for col_num in range(startcol_descricao + 1, endcol_descricao + 1):
                for idx in range(len(desc_produtos)):
                    cell = ws.cell(row=ini_prod + idx, column=col_num)
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_do_cod_e_qtde(self, writer, inicio_produtos, dados_tab):
        try:
            df_dados_um = pd.DataFrame(dados_tab, columns=['*COD. PROD.', 'QTDE'])
            codigo_int = {'*COD. PROD.': int}
            df_dados_um = df_dados_um.astype(codigo_int)
            qtde_float = {'QTDE': float}
            df_dados_um = df_dados_um.astype(qtde_float)
            df_dados_um.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=2, header=False,
                                 index=False)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_descricao(self, ws, writer, inicio_produtos, dados_tab):
        try:
            desc_produtos = [tabi[3] for tabi in dados_tab]
            df_descricao = pd.DataFrame({'DESCRIÇÃO DO PRODUTO': desc_produtos})
            df_descricao.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=4, header=False,
                                  index=False)

            ini_coluna_descr = 5
            fim_coluna_descr = 10

            for index, desc in enumerate(desc_produtos):
                num_linha = inicio_produtos + index

                ws.cell(row=num_linha, column=ini_coluna_descr).value = desc
                ws.cell(row=num_linha, column=ini_coluna_descr).alignment = Alignment(horizontal='center',
                                                                                      vertical='bottom',
                                                                                      wrap_text=True)

                if len(desc_produtos) > 1:
                    ws.merge_cells(start_row=num_linha,
                                   start_column=ini_coluna_descr,
                                   end_row=num_linha,
                                   end_column=fim_coluna_descr)

            for col_num in range(ini_coluna_descr + 1, fim_coluna_descr + 1):
                for idx in range(len(desc_produtos)):
                    cell = ws.cell(row=inicio_produtos + idx, column=col_num)
                    cell.alignment = Alignment(horizontal='center',
                                               vertical='bottom',
                                               wrap_text=True)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_da_um_ate_unid(self, writer, inicio_produtos, dados_tab):
        try:
            df = pd.DataFrame(dados_tab, columns=['UN', 'IND.', 'VLR UNIT.'])
            df.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=10, header=False, index=False)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_total(self, writer, inicio_produtos, dados_tab):
        try:
            df_dados_dois = pd.DataFrame(dados_tab, columns=['VLR TOTAL'])
            df_dados_dois.to_excel(writer, 'ORIGINAL', startrow=inicio_produtos - 1, startcol=13, header=False,
                                   index=False)

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaExpAlterar()
    tela.show()
    qt.exec_()
