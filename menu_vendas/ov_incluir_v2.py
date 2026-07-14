import sys
from banco_dados.conexao import conecta
from forms.tela_ov_incluir_v2 import *
from banco_dados.controle_erros import grava_erro_banco
from arquivos.chamar_arquivos import definir_caminho_arquivo
from comandos.tabelas import extrair_tabela, lanca_tabela, lanca_tabela_v2, layout_cabec_tab
from comandos.telas import tamanho_aplicacao, icone
from comandos.conversores import valores_para_float
from comandos.valores_padrao import custo_padrao_acinplas
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
from datetime import date, datetime, timedelta
import inspect
import os
import traceback


class TelaOvIncluirV2(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        frame = inspect.currentframe()
        if frame is None:
            raise RuntimeError("Não foi possível obter o frame atual.")
        nome_arquivo_com_caminho = inspect.getframeinfo(frame).filename
        self.nome_arquivo = os.path.basename(nome_arquivo_com_caminho)

        icone(self, "menu_vendas.png")
        tamanho_aplicacao(self)
        layout_cabec_tab(self.table_OV)

        self.combo_Cliente.activated.connect(self.lanca_exp_pendentes)

        self.table_Expedicao.viewport().installEventFilter(self)

        self.line_Frete.editingFinished.connect(self.mascara_frete)
        self.line_Desconto.editingFinished.connect(self.mascara_desconto)

        self.btn_ExcluirTudo.clicked.connect(self.excluir_tudo_ov)
        self.btn_ExcluirItem.clicked.connect(self.excluir_item_ov)

        self.btn_Limpar.clicked.connect(self.limpa_tudo)

        self.btn_Salvar.clicked.connect(self.verifica_salvamento)

        self.btn_Excel.clicked.connect(self.gera_excel)

        validator = QtGui.QRegExpValidator(QtCore.QRegExp(r'\d+'), self.line_Num_OV)
        self.line_Num_OV.setValidator(validator)

        self.definir_emissao()
        self.definir_combo_cliente_exp()

    def trata_excecao(self, nome_funcao, mensagem, arquivo, excecao):
        try:
            tb = traceback.extract_tb(excecao)
            num_linha_erro = tb[-1][1]

            traceback.print_exc()
            print(f'Houve um problema no arquivo: {arquivo} na função: "{nome_funcao}"\n{mensagem} {num_linha_erro}')
            self.mensagem_alerta(f'Houve um problema no arquivo:\n\n{arquivo}\n\n'
                                 f'Comunique o desenvolvedor sobre o problema descrito abaixo:\n\n'
                                 f'{nome_funcao}: {mensagem}')

            grava_erro_banco(nome_funcao, mensagem, arquivo, num_linha_erro)

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao_trat = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            tb = traceback.extract_tb(exc_traceback)
            num_linha_erro = tb[-1][1]
            print(f'Houve um problema no arquivo: {self.nome_arquivo} na função: "{nome_funcao_trat}"\n'
                  f'{e} {num_linha_erro}')
            grava_erro_banco(nome_funcao_trat, e, self.nome_arquivo, num_linha_erro)

    def mensagem_alerta(self, mensagem):
        try:
            alert = QMessageBox()
            alert.setIcon(QMessageBox.Warning)
            alert.setText(mensagem)
            alert.setWindowTitle("Atenção")
            alert.setStandardButtons(QMessageBox.Ok)
            alert.exec_()

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_emissao(self):
        try:
            data_hoje = date.today()
            self.date_Emissao.setDate(data_hoje)

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def definir_combo_cliente_exp(self):
        try:
            self.combo_Cliente.clear()

            nova_lista = [""]

            cursor = conecta.cursor()
            cursor.execute("""
                SELECT DISTINCT
                    c.id,
                    c.razao
                FROM EXPEDICAO_PRODUTO ep
                INNER JOIN PEDIDOINTERNO pi
                    ON pi.id = ep.num_origem
                INNER JOIN CLIENTES c
                    ON c.id = pi.id_cliente
                WHERE ep.tipo_origem = 'PI'
                ORDER BY c.razao
            """)

            for ides, descr in cursor.fetchall():
                nova_lista.append(f"{ides} - {descr}")

            self.combo_Cliente.addItems(nova_lista)

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def lanca_exp_pendentes(self):
        try:
            self.table_Expedicao.setRowCount(0)

            tabela_nova = []

            cliente = self.combo_Cliente.currentText()

            if cliente:
                clientetete = cliente.find(" - ")
                id_cliente = cliente[:clientetete]

                cursor = conecta.cursor()
                cursor.execute(f"""
                                SELECT ep.TIPO_ORIGEM, ep.NUM_ORIGEM, prod.codigo, prod.descricao, 
                                COALESCE(prod.obs, ''), prod.unidade, ep.qtde, ep.LOCAL_EXPEDICAO
                                FROM EXPEDICAO_PRODUTO ep
                                INNER JOIN EXPEDICAO exp ON exp.id = ep.ID_EXPEDICAO
                                INNER JOIN CLIENTES c ON c.id = exp.id_cliente
                                INNER JOIN produto prod ON ep.id_produto = prod.id
                                WHERE ep.tipo_origem = 'PI' 
                                and exp.id_cliente = {id_cliente}
                                ORDER BY c.razao
                            """)
                dados_interno = cursor.fetchall()

                if dados_interno:
                    for i in dados_interno:
                        tipo, num_ped, cod, descr, ref, um, qtde, local = i

                        origem = f"{tipo} {num_ped}"

                        preco = self.consulta_custo_produto(cod)

                        dados = (local, cod, descr, ref, um, qtde, preco, origem)

                        tabela_nova.append(dados)

            if tabela_nova:
                lista_de_listas_ordenada = sorted(tabela_nova, key=lambda x: x[2])
                lanca_tabela_v2(self.table_Expedicao, lista_de_listas_ordenada)

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def consulta_custo_produto(self, cod_prod):
        try:
            valor_s = 0

            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.id, prod.conjunto, prod.terceirizado, prod.custounitario, prod.custoestrutura "
                           f"FROM produto as prod "
                           f"where prod.codigo = {cod_prod};")
            dados_interno = cursor.fetchall()
            if dados_interno:
                id_prod, conj, terc, unit, estrut = dados_interno[0]

                if conj == 10:
                    cursor = conecta.cursor()
                    cursor.execute(f"UPDATE produto SET custoestrutura = '{0}' where id = {id_prod};")
                    conecta.commit()

                    custo_estrut_float = valores_para_float(estrut)
                    preco_estrut = custo_padrao_acinplas(custo_estrut_float)
                    valor_totau_dois = ("%.2f" % preco_estrut)
                    valor_s = str(valor_totau_dois)

                    estrutura = self.verifica_estrutura_problema(cod_prod)
                    for itens in estrutura:
                        conj_e, temp_e, terc_e, unit_e = itens

                        if conj_e == 10:
                            if temp_e or terc_e:
                                valor_s = str(valor_totau_dois)
                        else:
                            if unit_e:
                                valor_s = str(valor_totau_dois)
                else:
                    custo_tot_float = valores_para_float(unit)
                    preco = custo_padrao_acinplas(custo_tot_float)
                    valor_totau_dois = ("%.2f" % preco)
                    valor_s = str(valor_totau_dois)

            return valor_s

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def eventFilter(self, source, event):
        try:
            nome_tabela = self.table_Expedicao

            if (event.type() == QtCore.QEvent.MouseButtonDblClick and
                    event.buttons() == QtCore.Qt.LeftButton and
                    source is self.table_Expedicao.viewport()):
                self.label_Texto_Excel.setText(f'')

                item = nome_tabela.currentItem()

                extrai_recomendados = extrair_tabela(nome_tabela)
                item_selecionado = extrai_recomendados[item.row()]

                if item_selecionado:
                    self.manipula_dados_cliente(item_selecionado)

            return super(QMainWindow, self).eventFilter(source, event)

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def manipula_dados_cliente(self, item_selecionado):
        try:
            tabela_nova = []

            local, cod, descr, ref, um, qtde, preco, origem = item_selecionado

            origemtete = origem.find(" ")
            num_origem = origem[origemtete + 1:]

            cursor = conecta.cursor()
            cursor.execute(
                f"SELECT prod.id, prod.conjunto, COALESCE(prod.NCM, '') "
                f"FROM produto as prod "
                f"where prod.codigo = {cod};")
            dados_interno = cursor.fetchall()

            cursor = conecta.cursor()
            cursor.execute(f"SELECT COALESCE(ped.NUM_REQ_CLIENTE, ''), COALESCE(ped.obs, ''), ped.solicitante "
                           f"FROM pedidointerno as ped "
                           f"where ped.id = {num_origem};")
            dados_pedido = cursor.fetchall()

            if dados_interno and dados_pedido:
                id_prod, conj, ncm = dados_interno[0]

                req, obs, solic = dados_pedido[0]

                data_pr = date.today() + timedelta(days=7)
                data = f'{data_pr.day}/{data_pr.month}/{data_pr.year}'

                if preco:
                    if conj == 10:
                        qtde_float = valores_para_float(qtde)
                        preco_e = valores_para_float(preco)

                        total = qtde_float * preco_e
                        total_2 = ("%.2f" % total)

                        ncm = ncm.strip()

                        if (ncm == "8412.90.90" or ncm == "84129090" or ncm == "8412.9090"
                                or ncm == "8477.80.90" or ncm == "84778090" or ncm == "8477.8090"):
                            ipi = ""
                        else:
                            ipi = "3.25"


                        dados = [num_origem, req, cod, descr, ref, ncm, um, qtde, preco_e, ipi, total_2, data,
                                 obs, solic]

                        tabela_nova.append(dados)

                    else:
                        qtde_float = valores_para_float(qtde)
                        custo_t = valores_para_float(preco)

                        total = qtde_float * custo_t
                        total_dois = ("%.2f" % total)

                        dados = [num_origem, req, cod, descr, ref, ncm, um, qtde, custo_t, "", total_dois, data, obs, solic]

                        tabela_nova.append(dados)


            if tabela_nova:
                lanca_tabela(self.table_OV, tabela_nova)
                self.soma_totais()

        except Exception as e:
            nome_funcao = inspect.currentframe().f_code.co_name
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_estrutura_problema(self, codigo):
        try:
            cursor = conecta.cursor()
            cursor.execute(f"SELECT prod.id, prod.codigo, prod.descricao, prod.obs, prod.unidade, "
                           f"prod.conjunto, prod.tempo, prod.terceirizado, prod.custounitario, "
                           f"prod.custoestrutura, prod.id_versao "
                           f"FROM produto as prod "
                           f"LEFT JOIN tipomaterial as tip ON prod.tipomaterial = tip.id "
                           f"INNER JOIN conjuntos conj ON prod.conjunto = conj.id "
                           f"where prod.codigo = {codigo};")
            detalhes_pai = cursor.fetchall()
            (id_pai, c_pai, des_pai, ref_pai, um_pai, conj_pai, temp_pai, terc_pai, unit_pai, est_pai,
             id_estrut) = detalhes_pai[0]

            filhos = [(conj_pai, temp_pai, terc_pai, unit_pai)]

            if id_estrut:
                cursor = conecta.cursor()
                cursor.execute(f"SELECT prod.codigo, prod.descricao, COALESCE(prod.obs, '') as obs, prod.unidade "
                               f"FROM estrutura_produto as estprod "
                               f"INNER JOIN produto prod ON estprod.id_prod_filho = prod.id "
                               f"WHERE estprod.id_estrutura = {id_estrut};")
                dados_estrutura = cursor.fetchall()

                if dados_estrutura:
                    for prod in dados_estrutura:
                        cod_f, descr_f, ref_f, um_f = prod

                        filhos.extend(self.verifica_estrutura_problema(cod_f))

            return filhos

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def mascara_frete(self):
        try:
            self.soma_totais()

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def mascara_desconto(self):
        try:
            self.soma_totais()

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def soma_ipi(self):
        try:
            soma_com_ipi = 0.00

            extrai_produtos = extrair_tabela(self.table_OV)
            if extrai_produtos:
                for i in extrai_produtos:
                    qtde = i[7]
                    unit = i[8]
                    ipi = i[9]

                    qtde_float = valores_para_float(qtde)

                    unit_float = valores_para_float(unit)

                    if ipi:
                        ipi_float = valores_para_float(ipi)
                    else:
                        ipi_float = 0.00

                    ipi = qtde_float * unit_float * (ipi_float / 100)
                    soma_com_ipi += ipi

            total_ipi_final = str("%.2f" % soma_com_ipi)
            self.line_Total_Ipi.setText(total_ipi_final)

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def soma_totais(self):
        try:
            self.soma_ipi()

            soma_total_geral = 0.00
            soma_valores = 0.00
            soma_com_ipi = 0.00

            extrai_produtos = extrair_tabela(self.table_OV)
            if extrai_produtos:
                for i in extrai_produtos:
                    qtde = i[7]
                    unit = i[8]
                    ipi = i[9]
                    valor = i[10]

                    qtde_float = valores_para_float(qtde)

                    unit_float = valores_para_float(unit)

                    if ipi:
                        ipi_float = valores_para_float(ipi)
                    else:
                        ipi_float = 0.00

                    valor_com_ipi = qtde_float * ((unit_float * (ipi_float / 100)) + unit_float)

                    total_float = valores_para_float(valor)

                    soma_valores += total_float
                    soma_com_ipi += valor_com_ipi

            frete = self.line_Frete.text()
            if frete:
                valor_frete = round(valores_para_float(frete), 2)
            else:
                valor_frete = 0

            desconto = self.line_Desconto.text()
            if desconto:
                valor_desconto = round(valores_para_float(desconto), 2)
            else:
                valor_desconto = 0

            soma_total_geral += soma_com_ipi + valor_frete
            soma_total_geral -= valor_desconto

            total_mercadorias = str("%.2f" % soma_valores)
            self.line_Total_Merc.setText(total_mercadorias)

            total_geral = str("%.2f" % soma_total_geral)
            self.line_Total_Geral.setText(total_geral)

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tabelas(self):
        try:
            self.table_OV.setRowCount(0)
            self.table_PI_Aberto.setRowCount(0)

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def limpa_tudo(self):
        self.limpa_tabelas()
        self.line_Cliente.clear()
        self.line_Frete.clear()
        self.line_Desconto.clear()
        self.line_Total_Ipi.clear()
        self.line_Total_Merc.clear()
        self.line_Total_Geral.clear()
        self.line_Obs.clear()
        self.manipula_dados_pi()

    def excluir_tudo_ov(self):
        try:
            extrai_tabela_ov = extrair_tabela(self.table_OV)
            if not extrai_tabela_ov:
                self.mensagem_alerta(f'A tabela "Produtos Ordem de Venda" está vazia!')
            else:
                self.table_OV.setRowCount(0)
                self.soma_totais()

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def excluir_item_ov(self):
        try:
            extrai_recomendados = extrair_tabela(self.table_OV)
            if not extrai_recomendados:
                self.mensagem_alerta(f'A tabela "Produtos Ordem de Venda" está vazia!')
            else:
                linha_selecao = self.table_OV.currentRow()
                self.table_OV.removeRow(linha_selecao)
                self.soma_totais()

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def verifica_salvamento(self):
        try:
            num_ov = self.line_Num_OV.text()
            cliente = self.line_Cliente.text()
            tabela_produtos = extrair_tabela(self.table_OV)

            if not num_ov:
                self.mensagem_alerta('O campo "Nº OV" não pode estar vazio!')
            elif not cliente:
                self.mensagem_alerta('O campo "Cliente" não pode estar vazio!')
            elif not tabela_produtos:
                self.mensagem_alerta('A tabela "Produtos Ordem de Venda" não pode estar vazia"')
            else:
                num_ov = self.line_Num_OV.text()

                clientetete = cliente.find(" - ")
                id_cliente = cliente[:clientetete]

                cursor = conecta.cursor()
                cursor.execute(f"SELECT oc.cliente, oc.status, COALESCE(oc.descontos, '') as descon, "
                               f"COALESCE(oc.frete, '') as frete, "
                               f"COALESCE(oc.obs, '') as obs "
                               f"FROM ordemcompra as oc "
                               f"INNER JOIN clientes as cli ON oc.cliente = cli.id "
                               f"where oc.entradasaida = 'S' "
                               f"and oc.numero = {num_ov} "
                               f"and oc.cliente = {id_cliente};")
                dados_interno = cursor.fetchall()

                if dados_interno:
                    self.mensagem_alerta("Esta Ordem de Venda (OV) já foi adicionada!")
                else:
                    self.salvar_ov_novo()

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def salvar_ov_novo(self):
        try:
            num_ov = self.line_Num_OV.text()
            num_ov_int = int(num_ov)

            cliente = self.line_Cliente.text()
            clientetete = cliente.find(" - ")
            id_cliente = cliente[:clientetete]

            emissao_oc = self.date_Emissao.date()
            data_emi = emissao_oc.toString("yyyy-MM-dd")

            frete = self.line_Frete.text()
            frete_oc_float = valores_para_float(frete)

            desconto = self.line_Desconto.text()
            desconto_oc_float = valores_para_float(desconto)

            cursor = conecta.cursor()
            cursor.execute("select GEN_ID(GEN_ORDEMCOMPRA_ID,0) from rdb$database;")
            ultimo_oc0 = cursor.fetchall()
            ultimo_oc1 = ultimo_oc0[0]
            ultimo_oc = int(ultimo_oc1[0]) + 1

            cursor = conecta.cursor()
            cursor.execute(f"Insert into ordemcompra (ID, ENTRADASAIDA, NUMERO, DATA, STATUS, CLIENTE, "
                           f"LOCALESTOQUE, FRETE, DESCONTOS) "
                           f"values (GEN_ID(GEN_ORDEMCOMPRA_ID,1), "
                           f"'S', {num_ov_int}, '{data_emi}', 'A', '{id_cliente}', '1', {frete_oc_float}, "
                           f"{desconto_oc_float});")

            dados_alterados = extrair_tabela(self.table_OV)

            lista_pi = []

            for itens in dados_alterados:
                (num_pi, req_cli, cod_produto, descr, ref, ncm, um, qtde, unit, ipi, total, entr,
                 qtde_ent, obs, solic) = itens

                codigo_int = int(cod_produto)

                entrega_prod = datetime.strptime(entr, '%d/%m/%Y').date()

                qtde_item_float = valores_para_float(qtde)

                valor_unit_float = valores_para_float(unit)

                ipi_item_float = valores_para_float(ipi)

                cursor = conecta.cursor()
                cursor.execute(f"SELECT id, descricao FROM produto where codigo = {codigo_int};")
                dados_produto = cursor.fetchall()

                id_produto, descricao = dados_produto[0]

                num_pi_int = int(num_pi)

                cursor = conecta.cursor()
                cursor.execute(f"Insert into produtoordemcompra (ID, MESTRE, PRODUTO, QUANTIDADE, UNITARIO, "
                               f"IPI, DATAENTREGA, NUMERO, CODIGO, PRODUZIDO, ID_PEDIDO) "
                               f"values (GEN_ID(GEN_PRODUTOORDEMCOMPRA_ID,1), {ultimo_oc}, "
                               f"{id_produto}, {qtde_item_float}, {valor_unit_float}, {ipi_item_float}, "
                               f"'{entrega_prod}', {num_ov_int}, '{codigo_int}', 0.0, {num_pi_int});")

                cursor = conecta.cursor()
                cursor.execute(f"UPDATE produtopedidointerno SET STATUS = 'B' "
                               f"WHERE id_produto = {id_produto} "
                               f"and id_pedidointerno = {num_pi_int};")

                lista_pi.append(num_pi)

            conecta.commit()

            self.mensagem_alerta(f'Ordem de Compra foi lançada com sucesso!')

            if lista_pi:
                lista_sem_duplicatas = list(set(lista_pi))
                for iii in lista_sem_duplicatas:
                    numero_pi = iii[0]

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT ped.id, cli.razao, prod.codigo, prod.descricao, "
                                   f"COALESCE(prod.obs, '') as obs, "
                                   f"prod.unidade, prodint.qtde, prodint.data_previsao "
                                   f"FROM PRODUTOPEDIDOINTERNO as prodint "
                                   f"INNER JOIN produto as prod ON prodint.id_produto = prod.id "
                                   f"INNER JOIN pedidointerno as ped ON prodint.id_pedidointerno = ped.id "
                                   f"INNER JOIN clientes as cli ON ped.id_cliente = cli.id "
                                   f"where prodint.status = 'A' and ped.id = {numero_pi};")
                    dados_interno = cursor.fetchall()

                    if not dados_interno:
                        cursor = conecta.cursor()
                        cursor.execute(f"UPDATE pedidointerno SET STATUS = 'B' "
                                       f"WHERE id = {numero_pi};")

                        conecta.commit()

            self.limpa_tabelas()
            self.limpa_tudo()
            self.manipula_dados_pi()
            self.line_Num_OV.clear()

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)

    def gera_excel(self):
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Side, Alignment, Border, Font, PatternFill
            from unidecode import unidecode
            from sympy import frac
            from pathlib import Path

            cor_cinza = "A6A6A6"

            client = self.line_Cliente.text()
            posicao_cli = client.find("-") + 2
            cliente = client[posicao_cli:]

            dados_tabela = extrair_tabela(self.table_OV)

            if dados_tabela and cliente:
                data_hoje = date.today()
                data_certa = data_hoje.strftime("%d/%m/%Y")

                valor_t = self.line_Total_Geral.text()
                total_float = valores_para_float(valor_t)

                cliente_maiuscula = cliente.upper()
                cliente_certo = unidecode(cliente_maiuscula)

                d_um = []

                solicitante = ""
                observacao = ""

                for tabi in dados_tabela:
                    num_pi, req_cli, cod, desc, ref, ncm, um, qtde, unit, ipi, total, entr, qtde_ent, obs, solic = tabi

                    solicitante = solic
                    observacao = obs

                    if unit == 0.00:
                        unit_1_final = 0.00
                        total_1_final = 0.00
                        ipi_final = 0.00
                    else:
                        if "," in unit:
                            unit_1_com_ponto = unit.replace(',', '.')
                            unit_1_float = valores_para_float(unit_1_com_ponto)
                        else:
                            unit_1_float = valores_para_float(unit)

                        if "," in total:
                            total_1_com_ponto = total.replace(',', '.')
                            total_1_float = valores_para_float(total_1_com_ponto)
                        else:
                            total_1_float = valores_para_float(total)

                        unit_1_final = valores_para_float(unit_1_float)
                        total_1_final = valores_para_float(total_1_float)

                        if ipi == 0.00:
                            ipi_final = 0.00
                        else:
                            if "," in ipi:
                                ipi_com_ponto = ipi.replace(',', '.')
                                ipi_final = valores_para_float(ipi_com_ponto)
                            else:
                                ipi_final = valores_para_float(ipi)

                    if "," in qtde:
                        qtdezinha_com_ponto = qtde.replace(',', '.')
                        qtdezinha_float = valores_para_float(qtdezinha_com_ponto)
                    else:
                        qtdezinha_float = valores_para_float(qtde)

                    cursor = conecta.cursor()
                    cursor.execute(f"SELECT prod.id, COALESCE(prod.NCM, ''), conj.conjunto "
                                   f"FROM produto as prod "
                                   f"INNER JOIN conjuntos as conj ON conj.id = prod.conjunto "
                                   f"where codigo = {cod};")
                    dados_prodis = cursor.fetchall()
                    idis, ncm, conj = dados_prodis[0]

                    dados = (cod, desc, ref, ncm, conj, um, qtdezinha_float, unit_1_final, ipi_final, total_1_final,
                             entr, obs)
                    d_um.append(dados)

                df = pd.DataFrame(d_um, columns=['Código', 'Descrição', 'Referência', 'NCM', 'Classificação',
                                                 'UM', 'Qtde', 'unit', 'Ipi %', 'total', 'Data', 'Destino'])

                codigo_int = {'Código': int}
                df = df.astype(codigo_int)
                qtde_float = {'Qtde': float}
                df = df.astype(qtde_float)

                camino = os.path.join('..', 'arquivos', 'modelo excel', 'ov_Copia.xlsx')
                caminho_arquivo = definir_caminho_arquivo(camino)

                book = load_workbook(caminho_arquivo)

                desktop = Path.home() / "Desktop"
                desk_str = str(desktop)
                nome_req = f'\Requisição {cliente_certo}.xlsx'
                caminho = (desk_str + nome_req)

                writer = pd.ExcelWriter(caminho, engine='openpyxl')

                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

                linhas_frame = df.shape[0]
                colunas_frame = df.shape[1]

                linhas_certas = linhas_frame + 2 + 9
                colunas_certas = colunas_frame + 1

                ws = book.active

                inicia = 11
                rows = range(inicia, inicia + linhas_frame)
                columns = range(1, colunas_certas)

                ws.row_dimensions[linhas_certas + 2].height = 30
                ws.row_dimensions[linhas_certas + 4].height = 30

                for row in rows:
                    for col in columns:
                        ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center',
                                                                wrap_text=True)
                        ws.cell(row, col).border = Border(left=Side(border_style='thin', color='00000000'),
                                                          right=Side(border_style='thin', color='00000000'),
                                                          top=Side(border_style='thin', color='00000000'),
                                                          bottom=Side(border_style='thin', color='00000000'),
                                                          diagonal=Side(border_style='thick', color='00000000'),
                                                          diagonal_direction=0,
                                                          outline=Side(border_style='thin', color='00000000'),
                                                          vertical=Side(border_style='thin', color='00000000'),
                                                          horizontal=Side(border_style='thin', color='00000000'))

                ws.merge_cells(f'A8:E8')
                top_left_cell = ws[f'A8']
                c = ws[f'A8']
                c.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=14, bold=True)
                top_left_cell.value = 'Requisição ' + cliente_certo

                ws.merge_cells(f'F8:J8')
                top_left_cell = ws[f'F8']
                c = ws[f'F8']
                c.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=14, bold=True)
                top_left_cell.value = 'Emissão:  ' + data_certa

                ws.merge_cells(f'A{linhas_certas}:I{linhas_certas}')
                top_left_cell = ws[f'A{linhas_certas}']
                c = ws[f'A{linhas_certas}']
                c.alignment = Alignment(horizontal='center',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=True)
                top_left_cell.value = 'TOTAL:'

                estilo_total = PatternFill(start_color=cor_cinza, end_color=cor_cinza, fill_type="solid")
                ws[f'A{linhas_certas}'].fill = estilo_total
                ws[f'B{linhas_certas}'].fill = estilo_total
                ws[f'C{linhas_certas}'].fill = estilo_total
                ws[f'D{linhas_certas}'].fill = estilo_total
                ws[f'E{linhas_certas}'].fill = estilo_total
                ws[f'F{linhas_certas}'].fill = estilo_total
                ws[f'G{linhas_certas}'].fill = estilo_total
                ws[f'H{linhas_certas}'].fill = estilo_total
                ws[f'I{linhas_certas}'].fill = estilo_total
                ws[f'J{linhas_certas}'].fill = estilo_total
                ws[f'K{linhas_certas}'].fill = estilo_total
                ws[f'L{linhas_certas}'].fill = estilo_total

                decimais = frac(total_float)
                if decimais == 0:
                    if total_float == 0.00:
                        ws.merge_cells(f'J{linhas_certas}:J{linhas_certas}')
                        top_left_cell = ws[f'J{linhas_certas}']
                        c = ws[f'J{linhas_certas}']
                        c.alignment = Alignment(horizontal='center',
                                                vertical='center',
                                                text_rotation=0,
                                                wrap_text=False,
                                                shrink_to_fit=False,
                                                indent=0)
                        c.font = Font(size=12, bold=True)
                        c.number_format = 'R$ 0.00;[Red]-R$ 0.00'
                        top_left_cell.value = total_float
                    else:
                        ws.merge_cells(f'J{linhas_certas}:J{linhas_certas}')
                        top_left_cell = ws[f'J{linhas_certas}']
                        c = ws[f'J{linhas_certas}']
                        c.alignment = Alignment(horizontal='center',
                                                vertical='center',
                                                text_rotation=0,
                                                wrap_text=False,
                                                shrink_to_fit=False,
                                                indent=0)
                        c.font = Font(size=12, bold=True)
                        c.number_format = 'R$ #.##00;[Red]-R$ #.##00'
                        top_left_cell.value = total_float
                else:
                    ws.merge_cells(f'J{linhas_certas}:J{linhas_certas}')
                    top_left_cell = ws[f'J{linhas_certas}']
                    c = ws[f'J{linhas_certas}']
                    c.alignment = Alignment(horizontal='center',
                                            vertical='center',
                                            text_rotation=0,
                                            wrap_text=False,
                                            shrink_to_fit=False,
                                            indent=0)
                    c.font = Font(size=12, bold=True)
                    c.number_format = 'R$ #.##;[Red]-R$ #.##'
                    top_left_cell.value = total_float

                ws.merge_cells(f'B{linhas_certas + 2}:B{linhas_certas + 2}')
                top_left_cell = ws[f'B{linhas_certas + 2}']
                c = ws[f'B{linhas_certas + 2}']
                c.alignment = Alignment(horizontal='right',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=True)
                top_left_cell.value = "Máquina:  "

                ws.merge_cells(f'C{linhas_certas + 2}:H{linhas_certas + 2}')
                top_left_cell = ws[f'C{linhas_certas + 2}']
                c = ws[f'C{linhas_certas + 2}']
                c.alignment = Alignment(horizontal='left',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=False)
                top_left_cell.value = observacao

                ws.merge_cells(f'K{linhas_certas + 2}:K{linhas_certas + 2}')
                top_left_cell = ws[f'K{linhas_certas + 2}']
                c = ws[f'K{linhas_certas + 2}']
                c.alignment = Alignment(horizontal='right',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=True)
                top_left_cell.value = "Cliente:  "

                ws.merge_cells(f'L{linhas_certas + 2}:L{linhas_certas + 2}')
                top_left_cell = ws[f'L{linhas_certas + 2}']
                c = ws[f'L{linhas_certas + 2}']
                c.alignment = Alignment(horizontal='left',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=False)
                top_left_cell.value = cliente_maiuscula

                ws.merge_cells(f'B{linhas_certas + 4}:B{linhas_certas + 4}')
                top_left_cell = ws[f'B{linhas_certas + 4}']
                c = ws[f'B{linhas_certas + 4}']
                c.alignment = Alignment(horizontal='right',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=True)
                top_left_cell.value = "Responsável:  "

                ws.merge_cells(f'C{linhas_certas + 4}:J{linhas_certas + 4}')
                top_left_cell = ws[f'C{linhas_certas + 4}']
                c = ws[f'C{linhas_certas + 4}']
                c.alignment = Alignment(horizontal='left',
                                        vertical='center',
                                        text_rotation=0,
                                        wrap_text=False,
                                        shrink_to_fit=False,
                                        indent=0)
                c.font = Font(size=12, bold=False)
                top_left_cell.value = solicitante

                df.to_excel(writer, 'Sheet1', startrow=10, startcol=0, header=False, index=False)

                writer.save()
                self.label_Texto_Excel.setText(f'Excel gerado com sucesso!')

        except Exception as e:
            frame = inspect.currentframe()
            nome_funcao = frame.f_code.co_name if frame else "desconhecida"
            exc_traceback = sys.exc_info()[2]
            self.trata_excecao(nome_funcao, str(e), self.nome_arquivo, exc_traceback)


if __name__ == '__main__':
    qt = QApplication(sys.argv)
    tela = TelaOvIncluirV2()
    tela.show()
    qt.exec_()
